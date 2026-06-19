# -*- coding: utf-8 -*-
"""Genera el CATALOGO DE FUNCIONES del proyecto (modo solo lectura del codigo).

Recorre todos los .py de src/, app/ y scripts/ y, por cada def, extrae:
modulo, funcion, utilidad, entradas, salida. No ejecuta ni importa codigo
de usuario: solo parsea con ast. Excluye carpetas .ipynb_checkpoints.
"""
import ast
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
while not (ROOT / 'src').exists():
    ROOT = ROOT.parent

CARPETAS = ['src', 'app', 'scripts']
SALIDA = ROOT / '_catalogo_funciones' / 'catalogo_funciones.xlsx'


def primera_frase(texto):
    """Devuelve la primera frase no vacia de un docstring."""
    if not texto:
        return None
    texto = texto.strip()
    for sep in ['\n\n', '. ', '\n']:
        if sep in texto:
            cand = texto.split(sep)[0].strip()
            if cand:
                return cand.rstrip('.').strip()
    return texto.rstrip('.').strip()


def deducir_utilidad(node):
    """Deduce de forma prudente la utilidad si no hay docstring."""
    # Si solo tiene return de algo simple, o pass
    cuerpo = [n for n in node.body if not isinstance(n, ast.Expr)]
    if not cuerpo:
        return '(verificar)'
    nombres_return = []
    for n in ast.walk(node):
        if isinstance(n, ast.Return) and n.value is not None:
            nombres_return.append(True)
    nombre = node.name
    # Heuristica suave por prefijo del nombre
    prefijos = {
        'get_': 'Obtiene/devuelve un valor',
        'set_': 'Establece un valor',
        'load': 'Carga datos',
        'cargar': 'Carga datos',
        'save': 'Guarda datos',
        'guardar': 'Guarda datos',
        'render': 'Renderiza/genera salida',
        'generar': 'Genera contenido',
        'crear': 'Crea un elemento',
        'validar': 'Valida datos',
        'verificar': 'Verifica una condicion',
        'comprobar': 'Comprueba una condicion',
        'plot_': 'Genera un grafico',
        'graficar': 'Genera un grafico',
        'format': 'Formatea un valor',
        'formatear': 'Formatea un valor',
        'is_': 'Comprueba una condicion (booleano)',
        'parse': 'Parsea/analiza datos',
        'build': 'Construye un objeto',
        'calc': 'Calcula un valor',
        'calcular': 'Calcula un valor',
    }
    for p, txt in prefijos.items():
        if nombre.lower().startswith(p):
            return txt + ' (verificar)'
    return '(verificar)'


def firma_parametros(node):
    """Reconstruye la lista de parametros con anotaciones y defaults."""
    a = node.args
    partes = []

    posonly = getattr(a, 'posonlyargs', [])
    args = posonly + a.args
    defaults = a.defaults
    n_sin_def = len(args) - len(defaults)

    def fmt(arg, default=None):
        s = arg.arg
        if arg.annotation is not None:
            try:
                s += ': ' + ast.unparse(arg.annotation)
            except Exception:
                pass
        if default is not None:
            try:
                s += '=' + ast.unparse(default)
            except Exception:
                s += '=...'
        return s

    for i, arg in enumerate(args):
        d = defaults[i - n_sin_def] if i >= n_sin_def else None
        partes.append(fmt(arg, d))
        if posonly and arg is posonly[-1]:
            partes.append('/')

    if a.vararg:
        partes.append('*' + fmt(a.vararg))
    elif a.kwonlyargs:
        partes.append('*')

    for arg, d in zip(a.kwonlyargs, a.kw_defaults):
        partes.append(fmt(arg, d))

    if a.kwarg:
        partes.append('**' + fmt(a.kwarg))

    return ', '.join(partes) if partes else '(sin parametros)'


def descripcion_salida(node):
    """Describe que devuelve la funcion (anotacion de retorno o heuristica)."""
    if node.returns is not None:
        try:
            return ast.unparse(node.returns)
        except Exception:
            pass
    tiene_return_valor = False
    tiene_yield = False
    for n in ast.walk(node):
        if isinstance(n, ast.Return) and n.value is not None:
            tiene_return_valor = True
        if isinstance(n, (ast.Yield, ast.YieldFrom)):
            tiene_yield = True
    if tiene_yield:
        return 'generador (yield)'
    if tiene_return_valor:
        return '(valor, verificar)'
    return 'None'


def es_metodo(node, padres):
    return any(isinstance(p, ast.ClassDef) for p in padres)


def recorrer(arbol):
    """Genera (node, clase) para cada def/async def, registrando la clase contenedora."""
    resultados = []

    def visit(node, clase):
        for hijo in ast.iter_child_nodes(node):
            if isinstance(hijo, (ast.FunctionDef, ast.AsyncFunctionDef)):
                resultados.append((hijo, clase))
                visit(hijo, clase)  # funciones anidadas
            elif isinstance(hijo, ast.ClassDef):
                visit(hijo, hijo.name)
            else:
                visit(hijo, clase)

    visit(arbol, None)
    return resultados


filas = []
ficheros_vistos = 0

for carpeta in CARPETAS:
    base = ROOT / carpeta
    if not base.exists():
        continue
    for py in sorted(base.rglob('*.py')):
        if '.ipynb_checkpoints' in py.parts:
            continue
        ficheros_vistos += 1
        rel = py.relative_to(ROOT).as_posix()
        try:
            codigo = py.read_text(encoding='utf-8')
        except UnicodeDecodeError:
            codigo = py.read_text(encoding='latin-1')
        try:
            arbol = ast.parse(codigo)
        except SyntaxError as e:
            filas.append([rel, '(ERROR DE PARSEO)', f'No se pudo parsear: {e}', '', '', '', '', ''])
            continue
        for node, clase in recorrer(arbol):
            nombre_simple = node.name
            nombre = node.name
            if clase:
                nombre = f'{clase}.{nombre}'
            doc = ast.get_docstring(node)
            utilidad = primera_frase(doc) or deducir_utilidad(node)
            entradas = firma_parametros(node)
            salida = descripcion_salida(node)
            linea = node.lineno
            es_async = 'si' if isinstance(node, ast.AsyncFunctionDef) else 'no'
            es_privado = 'si' if nombre_simple.startswith('_') else 'no'
            filas.append([rel, nombre, utilidad, entradas, salida,
                          linea, es_async, es_privado])

# ---- Construir Excel ----
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Catalogo'

cabeceras = ['modulo', 'funcion', 'utilidad', 'entradas', 'salida',
             'linea', 'async', 'privado']
ws.append(cabeceras)

fill = PatternFill('solid', fgColor='1F4E79')
for c in range(1, len(cabeceras) + 1):
    celda = ws.cell(row=1, column=c)
    celda.font = Font(bold=True, color='FFFFFF')
    celda.fill = fill
    celda.alignment = Alignment(vertical='center', horizontal='left')

for fila in filas:
    ws.append(fila)

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

anchos = {'A': 42, 'B': 34, 'C': 60, 'D': 50, 'E': 24,
          'F': 8, 'G': 8, 'H': 9}
for col, w in anchos.items():
    ws.column_dimensions[col].width = w

for fila in ws.iter_rows(min_row=2):
    for celda in fila:
        celda.alignment = Alignment(vertical='top', wrap_text=True)

# ---- Hoja resumen ----
from collections import Counter
conteo = Counter(f[0] for f in filas)
ws2 = wb.create_sheet('Resumen')
ws2.append(['modulo', 'n_funciones'])
for c in range(1, 3):
    celda = ws2.cell(row=1, column=c)
    celda.font = Font(bold=True, color='FFFFFF')
    celda.fill = fill
for modulo, n in sorted(conteo.items()):
    ws2.append([modulo, n])
ws2.append(['TOTAL', len(filas)])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
ws2.cell(row=ws2.max_row, column=2).font = Font(bold=True)
ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 14
ws2.freeze_panes = 'A2'

SALIDA.parent.mkdir(parents=True, exist_ok=True)
wb.save(SALIDA)

print(f'Ficheros .py procesados: {ficheros_vistos}')
print(f'Funciones catalogadas: {len(filas)}')
print(f'Excel guardado en: {SALIDA.relative_to(ROOT).as_posix()}')
print('--- Funciones por modulo ---')
for modulo, n in sorted(conteo.items()):
    print(f'{n:4d}  {modulo}')
