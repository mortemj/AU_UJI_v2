"""
Exportadores del motor de linaje: JSON (OpenLineage), Markdown, XLSX y HTML.

Mismo estilo y mismo esquema para todas las fases. Formato numerico espanol
(punto de miles, coma decimal). Intenta usar la funcion del proyecto
`formato_numero_es`; si no esta disponible, usa un equivalente local.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING
import json

if TYPE_CHECKING:
    from .core import Linaje


# --------------------------------------------------------------------------- #
# Formato numerico espanol (reutiliza el del proyecto si existe)
# --------------------------------------------------------------------------- #
try:
    from utils.formatters import formato_numero_es as _fmt   # estandar del proyecto
except Exception:                                            # equivalente local
    def _fmt(numero, decimales: int = 0) -> str:
        if numero is None:
            return "-"
        s = f"{int(numero):,}" if decimales == 0 else f"{numero:,.{decimales}f}"
        return s.replace(",", "@").replace(".", ",").replace("@", ".")


def _delta_txt(d):
    """Texto del delta con signo y formato espanol."""
    if d is None:
        return "-"
    if d == 0:
        return "0 (sin cambio)"
    signo = "+" if d > 0 else "−"
    return f"{signo}{_fmt(abs(d))}"


# --------------------------------------------------------------------------- #
# 1) JSON  (esquema OpenLineage: job / run / inputs / outputs + facets)
# --------------------------------------------------------------------------- #
def a_json(lin: "Linaje", ruta) -> Path:
    ruta = Path(ruta)
    eventos = []
    for p in lin.pasos:
        sal = lin.datasets.get(p.salida)
        inputs = []
        for e in p.entradas:
            ds = lin.datasets.get(e)
            if ds is None:
                continue
            inputs.append({
                "name": ds.nombre,
                "facets": {
                    "schema": {"fields": ds.columnas, "count": ds.n_cols},
                    "stats": {"rowCount": ds.n_filas},
                    "source": {"file": ds.fichero, "origen": ds.procedencia},
                },
            })
        outputs = []
        if sal is not None:
            outputs.append({
                "name": sal.nombre,
                "facets": {
                    "schema": {"fields": sal.columnas, "count": sal.n_cols},
                    "stats": {"rowCount": sal.n_filas},
                    "source": {"file": sal.fichero, "origen": sal.procedencia},
                },
            })
        eventos.append({
            "eventType": "COMPLETE",
            "job": {"namespace": f"tfm.{lin.fase}", "name": p.id,
                    "facets": {"nivel": p.nivel, "notebook": p.notebook,
                               "funcion": p.funcion, "operacion": p.operacion}},
            "inputs": inputs,
            "outputs": outputs,
            "facets": {  # facet propio de diffs (extension del estandar)
                "diff": {
                    "deltaFilas": p.delta_filas(lin.datasets),
                    "motivoFilas": p.motivo_filas,
                    "colsAnadidas": p.cols_anadidas(lin.datasets),
                    "colsEliminadas": p.cols_eliminadas(lin.datasets),
                    "motivoCols": p.motivo_cols,
                    "dtypesCambiados": p.dtypes_cambiados,
                    "nota": p.nota,
                },
            },
        })
    doc = {
        "_schema": "OpenLineage-compatible (extendido con facet 'diff')",
        "fase": lin.fase, "nivel": lin.nivel, "titulo": lin.titulo,
        "generado": lin.generado, "eventos": eventos,
    }
    ruta.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    return ruta


# --------------------------------------------------------------------------- #
# 2) Markdown  (tabla lista para chuleta / anexos)
# --------------------------------------------------------------------------- #
def a_markdown(lin: "Linaje", ruta) -> Path:
    ruta = Path(ruta)
    L = [f"# {lin.titulo}", "",
         f"**Fase:** {lin.fase}  |  **Nivel:** {lin.nivel}  |  "
         f"**Transiciones:** {len(lin.pasos)}", "",
         "| # | Notebook | Función | Operación | Entrada | Salida | "
         "Filas ent. | Filas sal. | Δ filas | Motivo filas | Cols + | Cols − |",
         "|---|----------|---------|-----------|---------|--------|"
         "-----------:|-----------:|--------:|--------------|--------|--------|"]
    for i, p in enumerate(lin.pasos, 1):
        ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
        sal = lin.datasets.get(p.salida)
        fe = _fmt(ent.n_filas) if ent and ent.n_filas is not None else "-"
        fs = _fmt(sal.n_filas) if sal and sal.n_filas is not None else "-"
        ca = ", ".join(p.cols_anadidas(lin.datasets)) or "-"
        cd = ", ".join(p.cols_eliminadas(lin.datasets)) or "-"
        L.append(f"| {i} | {p.notebook} | {p.funcion or '-'} | {p.operacion} | "
                 f"{p.entradas[0] if p.entradas else '-'} | {p.salida} | {fe} | {fs} | "
                 f"{_delta_txt(p.delta_filas(lin.datasets))} | {p.motivo_filas} | {ca} | {cd} |")
    ruta.write_text("\n".join(L), encoding="utf-8")
    return ruta


# --------------------------------------------------------------------------- #
# 3) XLSX  (3 hojas: Resumen, Columnas, Filas)
# --------------------------------------------------------------------------- #
def a_xlsx(lin: "Linaje", ruta) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    ruta = Path(ruta)
    wb = Workbook()

    azul = PatternFill("solid", fgColor="2B6CB0")
    gris = PatternFill("solid", fgColor="EDF2F7")
    blanco = Font(color="FFFFFF", bold=True)
    borde = Border(*[Side(style="thin", color="CBD5E0")] * 4)

    def cabecera(ws, cols):
        for j, c in enumerate(cols, 1):
            cel = ws.cell(1, j, c)
            cel.fill = azul; cel.font = blanco
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cel.border = borde
        ws.row_dimensions[1].height = 30

    # -- Hoja Resumen --
    ws = wb.active; ws.title = "Resumen"
    cabecera(ws, ["#", "Notebook", "Función", "Operación", "Entrada", "Salida",
                  "Filas ent.", "Filas sal.", "Δ filas", "Motivo filas",
                  "Cols ent.", "Cols sal.", "Cols +", "Cols −"])
    for i, p in enumerate(lin.pasos, 1):
        ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
        sal = lin.datasets.get(p.salida)
        fila = [i, p.notebook, p.funcion or "-", p.operacion,
                p.entradas[0] if p.entradas else "-", p.salida,
                ent.n_filas if ent else None, sal.n_filas if sal else None,
                p.delta_filas(lin.datasets), p.motivo_filas,
                ent.n_cols if ent else None, sal.n_cols if sal else None,
                len(p.cols_anadidas(lin.datasets)), len(p.cols_eliminadas(lin.datasets))]
        for j, v in enumerate(fila, 1):
            cel = ws.cell(i + 1, j, v); cel.border = borde
            if j in (7, 8, 11, 12) and isinstance(v, int):
                cel.number_format = '#,##0'                      # miles (Excel ES)
            if i % 2 == 0:
                cel.fill = gris

    # -- Hoja Columnas --
    ws2 = wb.create_sheet("Columnas")
    cabecera(ws2, ["#", "Notebook", "Salida", "Tipo cambio", "Columna", "Motivo"])
    r = 2
    for i, p in enumerate(lin.pasos, 1):
        for c in p.cols_anadidas(lin.datasets):
            for j, v in enumerate([i, p.notebook, p.salida, "AÑADIDA", c, p.motivo_cols], 1):
                ws2.cell(r, j, v).border = borde
            r += 1
        for c in p.cols_eliminadas(lin.datasets):
            for j, v in enumerate([i, p.notebook, p.salida, "ELIMINADA", c, p.motivo_cols], 1):
                ws2.cell(r, j, v).border = borde
            r += 1

    # -- Hoja Filas (cascada) --
    ws3 = wb.create_sheet("Filas")
    cabecera(ws3, ["#", "Notebook", "Operación", "Filas ent.", "Filas sal.",
                   "Δ filas", "Motivo"])
    for i, p in enumerate(lin.pasos, 1):
        ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
        sal = lin.datasets.get(p.salida)
        fila = [i, p.notebook, p.operacion, ent.n_filas if ent else None,
                sal.n_filas if sal else None, p.delta_filas(lin.datasets), p.motivo_filas]
        for j, v in enumerate(fila, 1):
            cel = ws3.cell(i + 1, j, v); cel.border = borde
            if j in (4, 5, 6) and isinstance(v, int):
                cel.number_format = '#,##0'

    # anchos
    for ws_ in (ws, ws2, ws3):
        for col in ws_.columns:
            largo = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws_.column_dimensions[col[0].column_letter].width = min(max(largo + 2, 10), 48)

    wb.save(ruta)
    return ruta


# --------------------------------------------------------------------------- #
# 4) HTML  (grafo SVG offline + tablas de diff, estilo del proyecto)
# --------------------------------------------------------------------------- #
_CSS = """
 body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
   margin:0;padding:24px;color:#2d3748;background:#f7fafc}
 .wrap{max-width:1180px;margin:0 auto}
 h1{font-size:22px;margin:0 0 4px} .sub{color:#718096;margin:0 0 18px}
 .panel{background:white;border:1px solid #e2e8f0;border-radius:12px;padding:18px;margin:16px 0}
 table{border-collapse:collapse;width:100%;font-size:13px}
 th,td{border:1px solid #e2e8f0;padding:7px 9px;text-align:left;vertical-align:top}
 th{background:#2b6cb0;color:white}
 tr:nth-child(even) td{background:#f7fafc}
 code{background:#edf2f7;padding:1px 5px;border-radius:4px;font-size:12px}
 .toggle{display:inline-flex;border:1px solid #cbd5e0;border-radius:8px;overflow:hidden;margin:0 0 8px}
 .toggle button{border:0;background:white;color:#2b6cb0;padding:8px 16px;font-size:14px;
   cursor:pointer;font-weight:600}
 .toggle button.on{background:#2b6cb0;color:white}
 .descargas a{display:inline-block;margin-right:10px;font-size:13px;color:#2b6cb0;
   text-decoration:none;border:1px solid #cbd5e0;border-radius:6px;padding:5px 10px}
 .hidden{display:none}
"""


def _bloque(lin: "Linaje") -> tuple[str, str]:
    """Devuelve (svg_html, filas_tabla_html) para una fase a un nivel dado."""
    # Guarda: si no hay transiciones, no intentamos dibujar (evita max() vacío).
    if not lin.pasos:
        aviso = ('<p style="background:#fff5f5;border-left:4px solid #c53030;'
                 'padding:10px 14px;border-radius:6px">Sin transiciones registradas '
                 'en este nivel. Suele significar que no se encontraron los parquet de '
                 'entrada (revisa que <code>data/01_interim/</code> contenga las tablas '
                 'limpias).</p>')
        return aviso, ""
    # --- grafo SVG: nodos en columnas por orden de aparicion ---
    orden, vistos = [], set()
    for p in lin.pasos:
        for n in p.entradas + [p.salida]:
            if n not in vistos:
                vistos.add(n); orden.append(n)
    # nivel x de cada nodo = profundidad topologica simple
    nivel_x = {n: 0 for n in orden}
    for p in lin.pasos:
        sal = p.salida
        base = max((nivel_x[e] for e in p.entradas if e in nivel_x), default=-1)
        nivel_x[sal] = max(nivel_x.get(sal, 0), base + 1)
    cols: dict[int, list[str]] = {}
    for n, x in nivel_x.items():
        cols.setdefault(x, []).append(n)

    AN, AL = 210, 56          # ancho/alto nodo
    GX, GY = 110, 26          # gap horizontal/vertical
    W = (max(cols) + 1) * (AN + GX) + GX
    H = max(len(v) for v in cols.values()) * (AL + GY) + GY + 40
    pos = {}
    for x, nodos in cols.items():
        for k, n in enumerate(nodos):
            pos[n] = (GX + x * (AN + GX), 40 + k * (AL + GY))

    def color(n):
        ds = lin.datasets.get(n)
        if ds and ds.tipo == "excel":   return "#ed8936"
        if n == lin.pasos[-1].salida:   return "#805ad5"
        return "#3182ce"

    svg = [f'<svg viewBox="0 0 {W} {H}" class="linaje-svg" '
           f'xmlns="http://www.w3.org/2000/svg" style="max-width:100%;height:auto">']
    # aristas
    for p in lin.pasos:
        x2, y2 = pos[p.salida]
        for e in p.entradas:
            if e not in pos:
                continue
            x1, y1 = pos[e]
            mx = (x1 + AN + x2) / 2
            d = p.delta_filas(lin.datasets)
            etq = _delta_txt(d) if d not in (None, 0) else ""
            svg.append(f'<path d="M{x1+AN},{y1+AL/2} C{mx},{y1+AL/2} {mx},{y2+AL/2} '
                       f'{x2},{y2+AL/2}" fill="none" stroke="#a0aec0" stroke-width="1.6"/>')
            if etq:
                svg.append(f'<text x="{mx}" y="{(y1+y2)/2+AL/2-4}" font-size="11" '
                           f'fill="#c53030" text-anchor="middle">{etq}</text>')
    # nodos
    for n, (x, y) in pos.items():
        ds = lin.datasets.get(n)
        sub = (f'{_fmt(ds.n_filas)} filas · {ds.n_cols} cols'
               if ds and ds.n_filas is not None else "")
        svg.append(
            f'<g><rect x="{x}" y="{y}" width="{AN}" height="{AL}" rx="8" '
            f'fill="white" stroke="{color(n)}" stroke-width="2"/>'
            f'<rect x="{x}" y="{y}" width="6" height="{AL}" rx="3" fill="{color(n)}"/>'
            f'<text x="{x+16}" y="{y+22}" font-size="13" font-weight="700" '
            f'fill="#2d3748">{n}</text>'
            f'<text x="{x+16}" y="{y+40}" font-size="11" fill="#718096">{sub}</text></g>')
    svg.append("</svg>")

    # --- tabla de diffs ---
    filas = []
    for i, p in enumerate(lin.pasos, 1):
        ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
        sal = lin.datasets.get(p.salida)
        fe = _fmt(ent.n_filas) if ent and ent.n_filas is not None else "-"
        fs = _fmt(sal.n_filas) if sal and sal.n_filas is not None else "-"
        dlt = _delta_txt(p.delta_filas(lin.datasets))
        ca = ", ".join(p.cols_anadidas(lin.datasets)) or "—"
        cd = ", ".join(p.cols_eliminadas(lin.datasets)) or "—"
        filas.append(
            f"<tr><td>{i}</td><td><code>{p.notebook}</code></td>"
            f"<td>{p.funcion or '—'}</td><td>{p.operacion}</td>"
            f"<td>{p.entradas[0] if p.entradas else '—'} → {p.salida}</td>"
            f"<td style='text-align:right'>{fe} → {fs}</td>"
            f"<td style='text-align:center'><b>{dlt}</b></td>"
            f"<td>{p.motivo_filas}</td>"
            f"<td style='color:#2f855a'>{ca}</td><td style='color:#c53030'>{cd}</td></tr>")

    return "".join(svg), "".join(filas)


_CAB_TABLA = ("<tr><th>#</th><th>Notebook</th><th>Función</th><th>Operación</th>"
              "<th>Flujo</th><th>Filas</th><th>Δ</th><th>Motivo filas</th>"
              "<th>Cols +</th><th>Cols −</th></tr>")


def _aviso(*lins) -> str:
    decl = any(d.procedencia == "declarado" for lin in lins for d in lin.datasets.values())
    if not decl:
        return ""
    return ("<p style='background:#fffaf0;border-left:4px solid #ed8936;"
            "padding:8px 12px;font-size:13px;border-radius:6px'>Conteos marcados como "
            "<b>declarados</b> provienen del registro del proyecto. Al ejecutar el "
            "piloto en local contra los parquet reales se leen literalmente del fichero.</p>")


def _pagina(titulo, sub, ruta_style, cuerpo) -> str:
    return (f'<!DOCTYPE html><html lang="es"><head><meta charset="utf-8">'
            f'<meta name="viewport" content="width=device-width, initial-scale=1">'
            f'<title>{titulo}</title><link rel="stylesheet" href="{ruta_style}">'
            f'<style>{_CSS}</style></head><body><div class="wrap">'
            f'<h1>{titulo}</h1><p class="sub">{sub}</p>{cuerpo}</div></body></html>')


def a_html(lin: "Linaje", ruta, ruta_style: str = "../style.css") -> Path:
    """HTML de un solo nivel (compatibilidad)."""
    ruta = Path(ruta)
    svg, filas = _bloque(lin)
    cuerpo = (f'{_aviso(lin)}<div class="panel">{svg}</div>'
              f'<div class="panel"><table>{_CAB_TABLA}{filas}</table></div>')
    sub = f"Fase {lin.fase} · nivel <b>{lin.nivel}</b> · {len(lin.pasos)} transiciones"
    ruta.write_text(_pagina(lin.titulo, sub, ruta_style, cuerpo), encoding="utf-8")
    return ruta


def a_html_combinado(lin_nb: "Linaje", lin_fn: "Linaje", ruta,
                     ruta_style: str = "../style.css",
                     descargas: dict[str, str] | None = None,
                     titulo: str | None = None) -> Path:
    """Una sola pagina por fase con CONMUTADOR notebook/funcion + descargas."""
    ruta = Path(ruta)
    svg_nb, fil_nb = _bloque(lin_nb)
    svg_fn, fil_fn = _bloque(lin_fn)
    titulo = titulo or f"Linaje {lin_nb.fase} · Transformación"

    botones = ('<div class="toggle">'
               '<button id="b-nb" class="on" onclick="ver(\'nb\')">Por notebook</button>'
               '<button id="b-fn" onclick="ver(\'fn\')">Por función</button></div>')

    desc = ""
    if descargas:
        enlaces = "".join(f'<a href="{u}" download>⬇ {k.upper()}</a>'
                          for k, u in descargas.items())
        desc = f'<div class="descargas" style="margin:6px 0 4px">{enlaces}</div>'

    bloque_nb = (f'<div id="nivel-nb"><div class="panel">{svg_nb}</div>'
                 f'<div class="panel"><table>{_CAB_TABLA}{fil_nb}</table></div></div>')
    bloque_fn = (f'<div id="nivel-fn" class="hidden"><div class="panel">{svg_fn}</div>'
                 f'<div class="panel"><table>{_CAB_TABLA}{fil_fn}</table></div></div>')

    js = ("<script>function ver(n){"
          "document.getElementById('nivel-nb').classList.toggle('hidden',n!=='nb');"
          "document.getElementById('nivel-fn').classList.toggle('hidden',n!=='fn');"
          "document.getElementById('b-nb').classList.toggle('on',n==='nb');"
          "document.getElementById('b-fn').classList.toggle('on',n==='fn');}</script>")

    cuerpo = f"{desc}{botones}{_aviso(lin_nb, lin_fn)}{bloque_nb}{bloque_fn}{js}"
    sub = (f"Fase {lin_nb.fase} · {len(lin_nb.pasos)} transiciones (notebook) / "
           f"{len(lin_fn.pasos)} (función)")
    ruta.write_text(_pagina(titulo, sub, ruta_style, cuerpo), encoding="utf-8")
    return ruta


# --------------------------------------------------------------------------- #
# 5) Versiones COMBINADAS (un fichero por fase, ambos niveles dentro)
# --------------------------------------------------------------------------- #
def _eventos(lin: "Linaje") -> list:
    """Lista de eventos OpenLineage de una fase (usado por el JSON combinado)."""
    ev = []
    for p in lin.pasos:
        sal = lin.datasets.get(p.salida)
        ins = [{"name": d.nombre,
                "facets": {"schema": {"fields": d.columnas, "count": d.n_cols},
                           "stats": {"rowCount": d.n_filas},
                           "source": {"file": d.fichero, "origen": d.procedencia}}}
               for d in (lin.datasets.get(e) for e in p.entradas) if d]
        outs = ([{"name": sal.nombre,
                  "facets": {"schema": {"fields": sal.columnas, "count": sal.n_cols},
                             "stats": {"rowCount": sal.n_filas},
                             "source": {"file": sal.fichero, "origen": sal.procedencia}}}]
                if sal else [])
        ev.append({"eventType": "COMPLETE",
                   "job": {"namespace": f"tfm.{lin.fase}", "name": p.id,
                           "facets": {"nivel": p.nivel, "notebook": p.notebook,
                                      "funcion": p.funcion, "operacion": p.operacion}},
                   "inputs": ins, "outputs": outs,
                   "facets": {"diff": {"deltaFilas": p.delta_filas(lin.datasets),
                                       "motivoFilas": p.motivo_filas,
                                       "colsAnadidas": p.cols_anadidas(lin.datasets),
                                       "colsEliminadas": p.cols_eliminadas(lin.datasets),
                                       "motivoCols": p.motivo_cols}}})
    return ev


def a_json_combinado(lin_nb: "Linaje", lin_fn: "Linaje", ruta) -> Path:
    ruta = Path(ruta)
    doc = {"_schema": "OpenLineage-compatible (extendido con facet 'diff')",
           "fase": lin_nb.fase, "generado": lin_nb.generado,
           "niveles": {"notebook": _eventos(lin_nb), "funcion": _eventos(lin_fn)}}
    ruta.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")
    return ruta


def a_markdown_combinado(lin_nb: "Linaje", lin_fn: "Linaje", ruta) -> Path:
    ruta = Path(ruta)
    def tabla(lin):
        L = ["| # | Notebook | Función | Operación | Flujo | Filas | Δ filas | "
             "Motivo filas | Cols + |",
             "|---|----------|---------|-----------|-------|-------|--------:|"
             "--------------|--------|"]
        for i, p in enumerate(lin.pasos, 1):
            ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
            sal = lin.datasets.get(p.salida)
            fe = _fmt(ent.n_filas) if ent and ent.n_filas is not None else "-"
            fs = _fmt(sal.n_filas) if sal and sal.n_filas is not None else "-"
            ca = ", ".join(p.cols_anadidas(lin.datasets)) or "—"
            L.append(f"| {i} | {p.notebook} | {p.funcion or '—'} | {p.operacion} | "
                     f"{(p.entradas[0] if p.entradas else '—')} → {p.salida} | {fe}→{fs} | "
                     f"{_delta_txt(p.delta_filas(lin.datasets))} | {p.motivo_filas} | {ca} |")
        return "\n".join(L)
    md = (f"# Linaje {lin_nb.fase} · Transformación\n\n"
          f"## Nivel notebook ({len(lin_nb.pasos)} transiciones)\n\n{tabla(lin_nb)}\n\n"
          f"## Nivel función ({len(lin_fn.pasos)} transiciones)\n\n{tabla(lin_fn)}\n")
    ruta.write_text(md, encoding="utf-8")
    return ruta


def a_xlsx_combinado(lin_nb: "Linaje", lin_fn: "Linaje", ruta,
                     inventario: list[dict] | None = None) -> Path:
    """Un libro con Resumen/Columnas/Filas; columna 'Nivel' distingue ambos.
    Si se pasa `inventario`, añade una hoja 'Mapa' con los 14 notebooks."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    ruta = Path(ruta)
    wb = Workbook()
    azul = PatternFill("solid", fgColor="2B6CB0"); gris = PatternFill("solid", fgColor="EDF2F7")
    blanco = Font(color="FFFFFF", bold=True)
    borde = Border(*[Side(style="thin", color="CBD5E0")] * 4)

    def cab(ws, cols):
        for j, c in enumerate(cols, 1):
            cel = ws.cell(1, j, c); cel.fill = azul; cel.font = blanco
            cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cel.border = borde
        ws.row_dimensions[1].height = 30

    ws = wb.active; ws.title = "Resumen"
    cab(ws, ["Nivel", "#", "Notebook", "Función", "Operación", "Entrada", "Salida",
             "Filas ent.", "Filas sal.", "Δ filas", "Motivo filas", "Cols +", "Cols −"])
    r = 2
    for lin in (lin_nb, lin_fn):
        for i, p in enumerate(lin.pasos, 1):
            ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
            sal = lin.datasets.get(p.salida)
            fila = [lin.nivel, i, p.notebook, p.funcion or "-", p.operacion,
                    p.entradas[0] if p.entradas else "-", p.salida,
                    ent.n_filas if ent else None, sal.n_filas if sal else None,
                    p.delta_filas(lin.datasets), p.motivo_filas,
                    len(p.cols_anadidas(lin.datasets)), len(p.cols_eliminadas(lin.datasets))]
            for j, v in enumerate(fila, 1):
                cel = ws.cell(r, j, v); cel.border = borde
                if j in (8, 9) and isinstance(v, int):
                    cel.number_format = '#,##0'
                if lin.nivel == "funcion":
                    cel.fill = gris
            r += 1

    ws2 = wb.create_sheet("Columnas")
    cab(ws2, ["Nivel", "Notebook", "Salida", "Tipo cambio", "Columna", "Motivo"])
    r = 2
    for lin in (lin_nb, lin_fn):
        for p in lin.pasos:
            for tipo, lista in (("AÑADIDA", p.cols_anadidas(lin.datasets)), ("ELIMINADA", p.cols_eliminadas(lin.datasets))):
                for c in lista:
                    for j, v in enumerate([lin.nivel, p.notebook, p.salida, tipo, c, p.motivo_cols], 1):
                        ws2.cell(r, j, v).border = borde
                    r += 1

    ws3 = wb.create_sheet("Filas")
    cab(ws3, ["Nivel", "#", "Notebook", "Operación", "Filas ent.", "Filas sal.", "Δ filas", "Motivo"])
    r = 2
    for lin in (lin_nb, lin_fn):
        for i, p in enumerate(lin.pasos, 1):
            ent = lin.datasets.get(p.entradas[0]) if p.entradas else None
            sal = lin.datasets.get(p.salida)
            fila = [lin.nivel, i, p.notebook, p.operacion, ent.n_filas if ent else None,
                    sal.n_filas if sal else None, p.delta_filas(lin.datasets), p.motivo_filas]
            for j, v in enumerate(fila, 1):
                cel = ws3.cell(r, j, v); cel.border = borde
                if j in (5, 6, 7) and isinstance(v, int):
                    cel.number_format = '#,##0'
            r += 1

    hojas = [ws, ws2, ws3]
    if inventario:
        ws4 = wb.create_sheet("Mapa")
        cab(ws4, ["Notebook", "Rol (deducido)", "Título", "Funciones", "Merges", "¿Transforma?"])
        for i, nb in enumerate(inventario, 2):
            fila = [nb["notebook"], nb["rol"], nb.get("titulo", ""),
                    ", ".join(nb["funciones"]) or "—", nb["merges"],
                    "sí" if nb["transforma"] else "no"]
            for j, v in enumerate(fila, 1):
                ws4.cell(i, j, v).border = borde
        hojas.append(ws4)

    for ws_ in hojas:
        for col in ws_.columns:
            largo = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws_.column_dimensions[col[0].column_letter].width = min(max(largo + 2, 10), 48)
    wb.save(ruta)
    return ruta


# --------------------------------------------------------------------------- #
# 6) HTML con TRES vistas: por notebook · por función · mapa de notebooks
# --------------------------------------------------------------------------- #
def _tabla_funciones(lin_fn: "Linaje", funciones: list[dict]) -> str:
    """Vista 'por función': funciones reales (parseadas) + aporte de columnas."""
    fil_fn = []
    for f in funciones:
        fil_fn.append(f"<tr><td>{f['funcion']}</td><td><code>{f['notebook']}</code></td>"
                      f"<td>{f['rol']}</td></tr>")
    tabla_funcs = (f"<h3 style='font-size:15px;margin:4px 0'>Funciones reales en los notebooks</h3>"
                   f"<table><tr><th>Función</th><th>Notebook</th><th>Rol</th></tr>"
                   f"{''.join(fil_fn)}</table>")

    # aporte de columnas (de lin_fn)
    fil_ap = []
    for i, p in enumerate(lin_fn.pasos, 1):
        ca = ", ".join(p.cols_anadidas(lin_fn.datasets)) or "—"
        fil_ap.append(f"<tr><td>{p.entradas[0] if p.entradas else '—'}</td>"
                      f"<td><code>{p.notebook}</code></td><td>{ca}</td></tr>")
    tabla_ap = (f"<h3 style='font-size:15px;margin:18px 0 4px'>Aporte de columnas a "
                f"<code>df_alumno</code> (derivado del esquema real)</h3>"
                f"<table><tr><th>Tabla</th><th>Notebook</th><th>Columnas que aporta</th></tr>"
                f"{''.join(fil_ap)}</table>")
    return f'<div class="panel">{tabla_funcs}{tabla_ap}</div>'



def _svg_mapa(inventario: list[dict]) -> str:
    """Diagrama por carriles (swimlanes): un carril por rol, notebooks en orden."""
    # rol -> (etiqueta de carril, color)
    def carril(rol):
        if rol.startswith("transformaci"): return ("Transformación (mueven datos)", "#3182ce")
        if "orquesta" in rol:              return ("Orquestación", "#718096")
        if "dashboard" in rol:             return ("Dashboard", "#d69e2e")
        if "grafo" in rol:                 return ("Grafos", "#805ad5")
        return ("Reportes", "#38a169")     # reporte / trazabilidad
    orden_carriles = ["Transformación (mueven datos)", "Orquestación", "Reportes", "Dashboard", "Grafos"]
    lanes = {k: [] for k in orden_carriles}
    color = {}
    for nb in sorted(inventario, key=lambda x: x["notebook"]):
        et, col = carril(nb["rol"]); lanes[et].append(nb); color[et] = col
    lanes = {k: v for k, v in lanes.items() if v}

    AN, AL, GX, GY = 168, 64, 26, 22
    LAB = 150                                  # ancho etiqueta de carril
    maxn = max(len(v) for v in lanes.values())
    W = LAB + maxn * (AN + GX) + GX
    H = len(lanes) * (AL + GY) + GY
    svg = [f'<svg viewBox="0 0 {W} {H}" xmlns="http://www.w3.org/2000/svg" style="max-width:100%;height:auto">']
    y = GY
    for et, nbs in lanes.items():
        col = color[et]
        svg.append(f'<rect x="0" y="{y-6}" width="{W}" height="{AL+12}" fill="{col}" opacity="0.06"/>')
        svg.append(f'<text x="10" y="{y+AL/2}" font-size="12" font-weight="700" fill="{col}">{et}</text>')
        x = LAB
        prev = None
        for nb in nbs:
            if nb["funciones"]:
                funcs = ", ".join(nb["funciones"])
            elif nb["merges"]:
                funcs = f'{nb["merges"]} merge'
            else:
                funcs = ""
            svg.append(
                f'<g><rect x="{x}" y="{y}" width="{AN}" height="{AL}" rx="9" fill="white" '
                f'stroke="{col}" stroke-width="2"/>'
                f'<rect x="{x}" y="{y}" width="6" height="{AL}" rx="3" fill="{col}"/>'
                f'<text x="{x+14}" y="{y+22}" font-size="11.5" font-weight="700" fill="#2d3748">{nb["notebook"]}</text>'
                f'<text x="{x+14}" y="{y+40}" font-size="9.5" fill="#718096">{funcs[:30]}</text></g>')
            # flecha de cadena dentro de transformación
            if et.startswith("Transformaci") and prev is not None:
                svg.append(f'<path d="M{prev},{y+AL/2} L{x},{y+AL/2}" stroke="{col}" '
                           f'stroke-width="1.4" marker-end="url(#fl)"/>')
            prev = x + AN
            x += AN + GX
        y += AL + GY
    svg.insert(1, '<defs><marker id="fl" markerWidth="8" markerHeight="8" refX="7" refY="3" '
                  'orient="auto"><path d="M0,0 L7,3 L0,6 Z" fill="#3182ce"/></marker></defs>')
    svg.append("</svg>")
    return "".join(svg)



_TIPO_CHIP = {
    "parquet": ("\U0001F4E6", "#3182ce"),
    "excel":   ("\U0001F4D7", "#2f855a"),
    "xlsx":    ("\U0001F4CA", "#2f855a"),
    "html":    ("\U0001F310", "#805ad5"),
    "png":     ("\U0001F5BC\uFE0F", "#d69e2e"),
    "csv":     ("\U0001F4C4", "#718096"),
}


def _chips(items):
    if not items:
        return '<span style="color:#a0aec0;font-size:12px">—</span>'
    out = []
    for tipo, nombre in items:
        ico, col = _TIPO_CHIP.get(tipo, ("\U0001F4C4", "#718096"))
        etq = nombre if nombre else tipo
        out.append(f'<span style="display:inline-block;background:{col}1a;color:{col};'
                   f'border:1px solid {col}55;border-radius:6px;padding:2px 7px;margin:2px;'
                   f'font-size:11px;white-space:nowrap">{ico} {etq}</span>')
    return "".join(out)


def _mapa_io(inventario):
    """Mapa entrada->salida por notebook, agrupado por rol. Tarjetas flexbox."""
    def grupo(rol):
        if rol.startswith("transformaci"): return ("Transformación (mueven datos)", "#3182ce")
        if "orquesta" in rol:              return ("Orquestación", "#718096")
        if "dashboard" in rol:             return ("Dashboard", "#d69e2e")
        if "grafo" in rol:                 return ("Grafos", "#805ad5")
        return ("Reportes", "#2f855a")
    orden = ["Transformación (mueven datos)", "Orquestación", "Reportes", "Dashboard", "Grafos"]
    grupos = {k: [] for k in orden}; colg = {}
    for nb in sorted(inventario, key=lambda x: x["notebook"]):
        et, col = grupo(nb["rol"]); grupos[et].append(nb); colg[et] = col
    grupos = {k: v for k, v in grupos.items() if v}

    html = []
    for et, nbs in grupos.items():
        col = colg[et]
        html.append(f'<h3 style="font-size:14px;margin:16px 0 6px;color:{col}">'
                    f'<span style="display:inline-block;width:10px;height:10px;border-radius:3px;'
                    f'background:{col};margin-right:6px"></span>{et}</h3>')
        for nb in nbs:
            cuerpo = ", ".join(nb["funciones"]) if nb["funciones"] else (
                     f'{nb["merges"]} merge' if nb["merges"] else nb["titulo"][:40])
            tarjeta = (
                '<div style="display:flex;align-items:center;gap:10px;background:white;'
                'border:1px solid #e2e8f0;border-left:4px solid ' + col + ';border-radius:10px;'
                'padding:10px 12px;margin:6px 0;flex-wrap:wrap">'
                '<div style="flex:1;min-width:170px">'
                '<div style="font-weight:700;font-size:12.5px;color:#2d3748">'
                + nb["notebook"] + '</div>'
                '<div style="font-size:10.5px;color:#a0aec0">' + (cuerpo or "") + '</div></div>'
                '<div style="flex:2;min-width:180px"><div style="font-size:10px;color:#a0aec0;'
                'text-transform:uppercase;letter-spacing:.04em">Lee</div>' + _chips(nb["lee"]) + '</div>'
                '<div style="font-size:18px;color:' + col + '">\u2192</div>'
                '<div style="flex:2;min-width:180px"><div style="font-size:10px;color:#a0aec0;'
                'text-transform:uppercase;letter-spacing:.04em">Produce</div>' + _chips(nb["produce"]) + '</div>'
                '</div>')
            html.append(tarjeta)
    return '<div class="panel">' + "".join(html) + '</div>'


def _tabla_mapa(inventario: list[dict]) -> str:
    """Vista 'mapa': una fila por notebook con su rol deducido."""
    filas = []
    for nb in inventario:
        emoji = ("\U0001F527" if nb["transforma"] else
                 "\U0001F4CA" if "dashboard" in nb["rol"] else
                 "\U0001F578\uFE0F" if "grafo" in nb["rol"] else
                 "\U0001F9ED" if "orquesta" in nb["rol"] else "\U0001F4C4")
        funcs = ", ".join(nb["funciones"]) or "—"
        merges = f"{nb['merges']} merge" if nb["merges"] else "—"
        transf = "sí" if nb["transforma"] else "no"
        filas.append(
            f"<tr><td><code>{nb['notebook']}</code></td>"
            f"<td>{emoji} {nb['rol']}</td><td>{nb['titulo']}</td>"
            f"<td>{funcs}</td><td style='text-align:center'>{merges}</td>"
            f"<td style='text-align:center'>{transf}</td></tr>")
    return (f'<div class="panel"><table>'
            f"<tr><th>Notebook</th><th>Rol (deducido)</th><th>Título</th>"
            f"<th>Funciones</th><th>Merges</th><th>¿Transforma?</th></tr>"
            f"{''.join(filas)}</table>"
            f"<p style='font-size:12px;color:#718096;margin-top:8px'>Rol deducido del "
            f"contenido de cada notebook (revisable).</p></div>")


def a_html_3vistas(lin_nb: "Linaje", lin_fn: "Linaje", inventario: list[dict],
                   funciones: list[dict], ruta, ruta_style: str = "../style.css",
                   descargas: dict[str, str] | None = None,
                   titulo: str | None = None) -> Path:
    ruta = Path(ruta)
    from .inventario import grafo_fase
    grafo = grafo_fase(inventario)
    svg_nb, fil_nb = _bloque(lin_nb)
    titulo = titulo or f"Linaje {lin_nb.fase} · Transformación"

    botones = ('<div class="toggle">'
               '<button id="b-nb" class="on" onclick="ver(\'nb\')">Por notebook</button>'
               '<button id="b-fn" onclick="ver(\'fn\')">Por función</button>'
               '<button id="b-mp" onclick="ver(\'mp\')">Mapa F1 (14 notebooks)</button>'
               '<button id="b-gr" onclick="ver(\'gr\')">Grafo completo</button></div>')

    desc = ""
    if descargas:
        enlaces = "".join(f'<a href="{u}" download>⬇ {k.upper()}</a>'
                          for k, u in descargas.items())
        desc = f'<div class="descargas" style="margin:6px 0 4px">{enlaces}</div>'

    v_nb = (f'<div id="v-nb"><div class="panel">{svg_nb}</div>'
            f'<div class="panel"><table>{_CAB_TABLA}{fil_nb}</table></div></div>')
    v_fn = f'<div id="v-fn" class="hidden">{_tabla_funciones(lin_fn, funciones)}</div>'
    v_mp = (f'<div id="v-mp" class="hidden">'
            f'<div class="panel">{_svg_mapa(inventario)}</div>'
            f'<h2 style="font-size:16px;margin:18px 0 0">Entradas → salidas por notebook</h2>'
            f'{_mapa_io(inventario)}{_tabla_mapa(inventario)}</div>')
    v_gr = f'<div id="v-gr" class="hidden">{_vista_grafo(grafo)}</div>'

    js = ("<script>function ver(n){"
          "for (const k of ['nb','fn','mp','gr']){"
          "document.getElementById('v-'+k).classList.toggle('hidden',k!==n);"
          "document.getElementById('b-'+k).classList.toggle('on',k===n);}}</script>")

    cuerpo = f"{desc}{botones}{_aviso(lin_nb, lin_fn)}{v_nb}{v_fn}{v_mp}{v_gr}{js}"
    sub = (f"Fase {lin_nb.fase} · {len(lin_nb.pasos)} transiciones · "
           f"{len(funciones)} funciones · {len(inventario)} notebooks")
    ruta.write_text(_pagina(titulo, sub, ruta_style, cuerpo), encoding="utf-8")
    return ruta


# --------------------------------------------------------------------------- #
# 7) Grafo completo (notebooks + ficheros): Mermaid (estático) + interactivo
# --------------------------------------------------------------------------- #
_CLASE_COLOR = {           # tipo -> (fill, stroke) para Mermaid y vis
    "notebook": ("#edf2f7", "#2d3748"),
    "parquet":  ("#ebf8ff", "#3182ce"),
    "html":     ("#faf5ff", "#805ad5"),
    "png":      ("#fffaf0", "#d69e2e"),
    "xlsx":     ("#f0fff4", "#2f855a"),
    "excel":    ("#f0fff4", "#2f855a"),
    "csv":      ("#f7fafc", "#718096"),
}


def _mermaid(grafo: dict) -> str:
    """Texto Mermaid (flowchart LR). Reproducible y pegable en la memoria."""
    L = ["flowchart LR"]
    for n in grafo["nodos"]:
        lab = n["label"].replace("{", "(").replace("}", ")").replace('"', "'")
        if n["clase"] == "notebook":
            L.append(f'  {n["id"]}["{lab}"]:::nb')
        else:
            L.append(f'  {n["id"]}[("{lab}")]:::{n["tipo"]}')
    for a in grafo["aristas"]:
        L.append(f'  {a["src"]} --> {a["dst"]}')
    for clase, (fill, stroke) in _CLASE_COLOR.items():
        key = "nb" if clase == "notebook" else clase
        L.append(f'  classDef {key} fill:{fill},stroke:{stroke},stroke-width:1px;')
    return "\n".join(L)


def _vis_datos(grafo: dict) -> tuple[str, str]:
    """(nodes_json, edges_json) para vis-network."""
    nodos = []
    for n in grafo["nodos"]:
        fill, stroke = _CLASE_COLOR.get(n["tipo"], ("#fff", "#718096"))
        forma = "box" if n["clase"] == "notebook" else "ellipse"
        nodos.append({"id": n["id"], "label": n["label"], "shape": forma,
                      "color": {"background": fill, "border": stroke},
                      "font": {"size": 13 if n["clase"] == "notebook" else 11}})
    aristas = [{"from": a["src"], "to": a["dst"], "arrows": "to",
                "color": {"color": "#cbd5e0"}} for a in grafo["aristas"]]
    return json.dumps(nodos, ensure_ascii=False), json.dumps(aristas, ensure_ascii=False)


def _vista_grafo(grafo: dict) -> str:
    """Sub-vista 'Grafo': conmutador Estático(Mermaid) / Interactivo(vis-network)."""
    mer = _mermaid(grafo)
    nodes_js, edges_js = _vis_datos(grafo)
    leyenda = "".join(
        f'<span style="margin-right:10px;font-size:12px"><span style="display:inline-block;'
        f'width:11px;height:11px;border-radius:3px;background:{c[0]};border:1px solid {c[1]};'
        f'margin-right:4px"></span>{t}</span>'
        for t, c in _CLASE_COLOR.items() if t != "excel")
    return f"""
<div class="toggle" style="margin-bottom:8px">
  <button id="g-est" class="on" onclick="vg('est')">Estático (Mermaid)</button>
  <button id="g-int" onclick="vg('int')">Interactivo</button>
</div>
<div style="margin:4px 0 8px">{leyenda}</div>
<div id="g-est-v" class="panel"><pre class="mermaid">{mer}</pre>
  <details style="margin-top:8px"><summary style="cursor:pointer;font-size:12px;color:#718096">
  Ver código Mermaid (pegable en la memoria)</summary>
  <pre style="background:#f7fafc;padding:10px;border-radius:6px;font-size:11px;overflow:auto">{mer}</pre>
  </details></div>
<div id="g-int-v" class="panel hidden">
  <p id="g-int-aviso" style="font-size:12px;color:#a0aec0">Cargando grafo interactivo… (requiere conexión)</p>
  <div id="red" style="height:560px;border:1px solid #e2e8f0;border-radius:8px"></div>
</div>
<script type="module">
  import mermaid from 'https://cdn.jsdelivr.net/npm/mermaid@11/dist/mermaid.esm.min.mjs';
  mermaid.initialize({{startOnLoad:true}});
</script>
<script src="https://unpkg.com/vis-network/standalone/umd/vis-network.min.js"></script>
<script>
  function vg(m){{
    document.getElementById('g-est-v').classList.toggle('hidden', m!=='est');
    document.getElementById('g-int-v').classList.toggle('hidden', m!=='int');
    document.getElementById('g-est').classList.toggle('on', m==='est');
    document.getElementById('g-int').classList.toggle('on', m==='int');
    if(m==='int' && !window._red){{ dibujarRed(); }}
  }}
  function dibujarRed(){{
    if(typeof vis==='undefined'){{ document.getElementById('g-int-aviso').textContent=
      'No se pudo cargar la librería (sin conexión). El grafo estático sí está disponible.'; return; }}
    const nodes=new vis.DataSet({nodes_js});
    const edges=new vis.DataSet({edges_js});
    window._red=new vis.Network(document.getElementById('red'), {{nodes,edges}}, {{
      layout:{{improvedLayout:true}}, physics:{{stabilization:true}},
      interaction:{{hover:true}}, nodes:{{borderWidth:1.5}} }});
    document.getElementById('g-int-aviso').style.display='none';
  }}
</script>"""
