# ============================================================================
# src/index_generator.py
# ============================================================================
#
# Genera y actualiza el index.html principal del proyecto.
#
# Uso desde cualquier notebook:
#   from src.index_generator import actualizar_index
#   actualizar_index()
#
# La función lee dinámicamente los datos disponibles (Excel, parquets, JSON
# de métricas del modelo) y regenera el index.html con KPIs actualizados.
#
# Si un dato no está disponible (fase no ejecutada), el KPI aparece
# translúcido con "Pendiente". Las tarjetas de Fase 5/6/7 se activan
# automáticamente cuando se detecta la señal correspondiente en disco:
#   - Fase 5: data/05_modelado/results/resultados_maestro.parquet
#   - Fase 6: data/06_evaluacion/metricas_modelo.json
#   - Fase 7: app/main.py
#
# NO tiene parámetros — todo se importa de src.config.
# ============================================================================

from datetime import datetime
from pathlib import Path


def actualizar_index(verbose=True):
    """
    Regenera docs/html/index.html con KPIs y secciones dinámicas.

    Lee los datos disponibles (Excel originales, parquets de cada fase,
    JSON de métricas del modelo) y construye el HTML con:
      - KPIs actualizados (período, expedientes, variables, abandono, AUC)
      - Tarjetas de Fase 5/6/7 activas según señales en disco
      - Sección de métricas del modelo ganador (si existe el JSON)
      - Aviso instructivo si faltan los Excel originales

    Parameters
    ----------
    verbose : bool
        Si True, imprime progreso por consola.

    Returns
    -------
    Path
        Ruta del index.html generado.
    """
    import pandas as pd
    import json as _json

    # --- Imports de config (todo centralizado) ---
    from src.config import (
        BASE_PATH,
        RUTA_RAW, RUTA_FEATURES, RUTA_HTML,
        RUTA_MODELADO, RUTA_EVALUACION,
        EXCEL_PRINCIPAL, EXCEL_PREINSCRIPCION,
        DATASET_FINAL_PARQUET,
        AUTORA, EMAIL_UOC, EMAIL_UJI, GITHUB_REPO,
        URL_APP_STREAMLIT,
        COLORES,
    )

    if verbose:
        print('🔄 Actualizando index.html...')

    # =================================================================
    # PASO 1: Datos del Excel con openpyxl (rápido, read_only)
    # =================================================================
    # Si los Excel no existen o no se pueden leer, n_tablas/n_registros
    # quedan en None — el HTML mostrará un aviso instructivo en vez de
    # inventar números (regla absoluta del proyecto: cero hardcodes).
    excel_disponible = False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL_PRINCIPAL, read_only=True)
        hojas = wb.sheetnames
        n_tablas = len(hojas)
        n_registros_raw = 0

        for hoja in hojas:
            ws = wb[hoja]
            n_filas = ws.max_row - 1 if ws.max_row else 0
            n_registros_raw += n_filas
        wb.close()

        if EXCEL_PREINSCRIPCION.exists():
            wb2 = load_workbook(EXCEL_PREINSCRIPCION, read_only=True)
            ws2 = wb2[wb2.sheetnames[0]]
            n_registros_raw += ws2.max_row - 1 if ws2.max_row else 0
            n_tablas += 1
            wb2.close()

        excel_disponible = True
    except Exception as e:
        if verbose:
            print(f'  \u26a0 Error leyendo Excel: {e}')
        hojas = []
        n_tablas = None
        n_registros_raw = None

    n_archivos_excel = len(list(RUTA_RAW.glob('*.xlsx'))) if RUTA_RAW.exists() else 0

    # =================================================================
    # PASO 2: Cargar metricas_modelo.json (Fases 5/6 — fuente de verdad)
    # =================================================================
    metricas_modelo = None
    ruta_metricas = RUTA_EVALUACION / 'metricas_modelo.json'
    if ruta_metricas.exists():
        try:
            with open(ruta_metricas, 'r', encoding='utf-8') as f:
                metricas_modelo = _json.load(f)
        except Exception as e:
            if verbose:
                print(f'  \u26a0 Error leyendo metricas_modelo.json: {e}')

    # =================================================================
    # PASO 3: KPIs dinámicos (prioriza JSON, fallback a parquets)
    # =================================================================
    kpi_expedientes = None
    kpi_periodo = None
    kpi_n_anios = None
    kpi_variables = None
    kpi_abandono = None
    kpi_auc = None

    # --- Prioridad 1: leer del JSON si existe ---
    if metricas_modelo:
        kpi_expedientes = metricas_modelo.get('n_alumnos_unicos')
        p_min = metricas_modelo.get('periodo_inicio')
        p_max = metricas_modelo.get('periodo_fin')
        if p_min and p_max:
            kpi_periodo = f'{p_min}-{p_max}'
            kpi_n_anios = p_max - p_min + 1
        kpi_variables = metricas_modelo.get('n_features')
        ta = metricas_modelo.get('tasa_abandono')
        if ta is not None:
            kpi_abandono = round(ta * 100, 1)
        kpi_auc = metricas_modelo.get('auc')

    # --- Prioridad 2: fallback al parquet de Fase 1 (si JSON no tenia datos) ---
    if (kpi_periodo is None or kpi_expedientes is None) and DATASET_FINAL_PARQUET.exists():
        df = pd.read_parquet(DATASET_FINAL_PARQUET)
        if kpi_periodo is None:
            for col in ['curso_aca', 'curso_aca_id']:
                if col in df.columns:
                    p_min = int(df[col].min())
                    p_max = int(df[col].max())
                    kpi_periodo = f'{p_min}-{p_max}'
                    kpi_n_anios = p_max - p_min + 1
                    break
        if kpi_expedientes is None:
            for col in ['per_id_ficticio', 'nip', 'alumno_id']:
                if col in df.columns:
                    kpi_expedientes = df[col].nunique()
                    break
        del df

    # --- Prioridad 3: fallback al parquet de Fase 3 (si aun faltan) ---
    ruta_feat = RUTA_FEATURES / 'df_expediente_features.parquet'
    if (kpi_variables is None or kpi_abandono is None) and ruta_feat.exists():
        df_feat = pd.read_parquet(ruta_feat)
        if kpi_variables is None:
            kpi_variables = len(df_feat.columns) - 1
        if kpi_abandono is None and 'abandono' in df_feat.columns:
            kpi_abandono = round((df_feat['abandono'] == 1).mean() * 100, 1)
        del df_feat

    # =================================================================
    # PASO 4: Detectar señales de fases hechas (sistema dinámico)
    # =================================================================
    # Fuente única de verdad: src/html/estado_proyecto.py contiene las 9
    # señales de las fases del proyecto (F0-F7 + AutoML). Aquí solo se
    # extraen los 3 booleanos que el resto de este fichero ya consume.
    # Si en el futuro cambia una ruta de fichero señal, se modifica una
    # vez en estado_proyecto.py — este fichero no requiere cambios.
    from src.html.estado_proyecto import detectar_estado_fases
    _estados = {e['id']: e['hecha'] for e in detectar_estado_fases()}
    fase5_done = _estados['fase5']
    fase6_done = _estados['fase6']
    fase7_done = _estados['fase7']

    # =================================================================
    # PASO 5: Construir HTML
    # =================================================================
    fecha = datetime.now().strftime('%d/%m/%Y')

    def fmt_es_num(n):
        """Formato número español: punto miles."""
        return f'{n:,}'.replace(',', '.')

    def fmt_es_dec(n, decimales=4):
        """Formato decimal español: coma decimal."""
        return f'{n:.{decimales}f}'.replace('.', ',')

    def kpi_html(valor, label, color, disponible=True):
        opacity = '' if disponible else ' opacity:0.35;'
        texto = valor if disponible else 'Pendiente'
        return (
            f'\n                <div class="kpi" style="{opacity}">'
            f'\n                    <div class="value" style="color:{color}">{texto}</div>'
            f'\n                    <div class="label">{label}</div>'
            f'\n                </div>'
        )

    # --- KPIs (5 ahora, antes 4) ---
    kpis_gen = ''
    kpis_gen += kpi_html(
        fmt_es_num(kpi_expedientes) if kpi_expedientes else '',
        'Expedientes únicos', COLORES.PRINCIPAL, kpi_expedientes is not None)
    kpis_gen += kpi_html(
        f'{kpi_periodo} ({kpi_n_anios} años)' if kpi_periodo else '',
        'Período', COLORES.OK, kpi_periodo is not None)
    kpis_gen += kpi_html(
        f'{kpi_variables}' if kpi_variables else '',
        'Variables del modelo', COLORES.ALERTA, kpi_variables is not None)
    kpis_gen += kpi_html(
        f'{fmt_es_dec(kpi_abandono, 1)}%' if kpi_abandono is not None else '',
        'Tasa abandono', COLORES.ERROR, kpi_abandono is not None)
    kpis_gen += kpi_html(
        fmt_es_dec(kpi_auc, 4) if kpi_auc is not None else '',
        'AUC del modelo', COLORES.PRINCIPAL, kpi_auc is not None)

    hay_pendientes = (kpi_expedientes is None or kpi_variables is None
                      or kpi_auc is None)
    aviso = ''
    if hay_pendientes:
        aviso = (
            '\n            <div style="background:#EBF8FF; padding:15px; border-radius:8px;'
            ' margin-top:15px; border-left:4px solid #3182ce; opacity:0.8;">'
            '\n                <strong>\u2139\ufe0f Nota:</strong> Los KPIs translúcidos se activarán'
            ' cuando ejecutes las fases correspondientes y vuelvas a ejecutar este notebook.'
            '\n            </div>'
        )

    hojas_texto = ', '.join(hojas + ['Preinscripción']) if hojas else 'No disponible'
    assets_rel = '../assets'

    # --- HTML completo ---
    H = []
    H.append('<!DOCTYPE html>')
    H.append('<html lang="es">')
    H.append('<head>')
    H.append('    <meta charset="UTF-8">')
    H.append('    <meta name="viewport" content="width=device-width, initial-scale=1.0">')
    H.append('    <title>TFM: Predicción de Abandono Universitario</title>')
    H.append('    <link rel="stylesheet" href="style.css">')
    H.append('</head>')
    H.append('<body>')
    H.append('    <header>')
    H.append('        <div class="header-content">')
    H.append(f'            <img src="{assets_rel}/logo_uoc.png" alt="UOC" class="header-logo" onerror="this.style.display=\'none\'">')
    H.append('            <div class="header-title">')
    H.append('                <h1>\U0001f393 TFM: Predicción de Abandono Universitario</h1>')
    H.append('                <p>Análisis y Modelado de Datos de Estudiantes UJI</p>')
    H.append('            </div>')
    H.append(f'            <img src="{assets_rel}/logo_uji.jpg" alt="UJI" class="header-logo" onerror="this.style.display=\'none\'">')
    H.append('        </div>')
    H.append('    </header>')
    H.append('')

    # --- Nav superior dinámica ---
    H.append('    <nav class="nav-fases">')
    H.append('        <a href="index.html" class="active">\U0001f3e0 Inicio</a>')
    H.append('        <a href="fase1/fase1_index.html">\U0001f4e5 Transformación</a>')
    H.append('        <a href="fase2/fase2_index.html">\U0001f4ca EDA Datos Originales</a>')
    H.append('        <a href="fase3/fase3_index.html">\U0001f527 Features</a>')
    H.append('        <a href="fase4/fase4_index.html">\U0001f52c EDA Final</a>')
    cls5 = '' if fase5_done else ' class="disabled"'
    cls6 = '' if fase6_done else ' class="disabled"'
    cls7 = '' if fase7_done else ' class="disabled"'
    H.append(f'        <a href="fase5/fase5_index.html"{cls5}>\U0001f916 Modelado</a>')
    H.append(f'        <a href="fase6/fase6_index.html"{cls6}>\U0001f4c8 Evaluación</a>')
    H.append(f'        <a href="fase7/fase7_index.html"{cls7}>\U0001f680 Aplicación</a>')
    H.append('        <a href="fase_automl/fase_automl_index.html">\u26a1 Pre-Modelado: AutoML</a>')
    H.append('    </nav>')
    H.append('')

    H.append('    <div class="container">')
    H.append('')
    H.append('        <section class="seccion">')
    H.append('            <h2>\U0001f4cb Estructura del Proyecto</h2>')
    H.append('            <p>Haz clic en una fase para explorar todos sus módulos.</p>')
    H.append('            <div class="grid-tarjetas">')

    # --- Tarjetas Fase 1-4 (siempre activas - son la base del proyecto) ---
    n_tablas_texto = f'{n_tablas} tablas' if n_tablas is not None else 'sus tablas'
    H.append('                <a href="fase1/fase1_index.html" class="tarjeta" style="border-left-color:#3182ce">')
    H.append('                    <h3 style="color:#3182ce">\U0001f4e5 Fase 1: Transformación</h3>')
    H.append(f'                    <p>Limpieza, unificación de {n_tablas_texto} y creación de df_alumno.</p>')
    H.append('                    <div class="link" style="color:#3182ce">Abrir fase \u2192</div>')
    H.append('                </a>')
    H.append('                <a href="fase2/fase2_index.html" class="tarjeta" style="border-left-color:#38a169">')
    H.append('                    <h3 style="color:#38a169">\U0001f4ca Fase 2: EDA Datos Originales</h3>')
    H.append('                    <p>Análisis exploratorio de los datos originales.</p>')
    H.append('                    <div class="link" style="color:#38a169">Abrir fase \u2192</div>')
    H.append('                </a>')
    H.append('                <a href="fase3/fase3_index.html" class="tarjeta" style="border-left-color:#805ad5">')
    H.append('                    <h3 style="color:#805ad5">\U0001f527 Fase 3: Feature Engineering</h3>')
    H.append('                    <p>Ingeniería de variables y preparación para modelado.</p>')
    H.append('                    <div class="link" style="color:#805ad5">Abrir fase \u2192</div>')
    H.append('                </a>')
    H.append('                <a href="fase4/fase4_index.html" class="tarjeta" style="border-left-color:#ed8936">')
    H.append('                    <h3 style="color:#ed8936">\U0001f52c Fase 4: EDA Final</h3>')
    H.append('                    <p>Análisis exploratorio del dataset final de features.</p>')
    H.append('                    <div class="link" style="color:#ed8936">Abrir fase \u2192</div>')
    H.append('                </a>')

    # --- Tarjetas Fase 5/6/7 - DINÁMICAS según señales en disco ---
    op5 = '' if fase5_done else '; opacity:0.6'
    link5 = 'Abrir fase \u2192' if fase5_done else 'Pendiente'
    H.append(f'                <a href="fase5/fase5_index.html" class="tarjeta" style="border-left-color:#e53e3e{op5}">')
    H.append('                    <h3 style="color:#e53e3e">\U0001f916 Fase 5: Modelado</h3>')
    H.append('                    <p>Entrenamiento de modelos de clasificación supervisada.</p>')
    H.append(f'                    <div class="link" style="color:#e53e3e">{link5}</div>')
    H.append('                </a>')

    op6 = '' if fase6_done else '; opacity:0.6'
    link6 = 'Abrir fase \u2192' if fase6_done else 'Pendiente'
    H.append(f'                <a href="fase6/fase6_index.html" class="tarjeta" style="border-left-color:#319795{op6}">')
    H.append('                    <h3 style="color:#319795">\U0001f4c8 Fase 6: Evaluación</h3>')
    H.append('                    <p>Comparativa de modelos, métricas e interpretabilidad.</p>')
    H.append(f'                    <div class="link" style="color:#319795">{link6}</div>')
    H.append('                </a>')

    op7 = '' if fase7_done else '; opacity:0.6'
    if fase7_done:
        # Tarjeta Fase 7 activa: descripción + link a la app desplegada
        descripcion_fase7 = (
            'Dashboard, vistas por perfil y documentación final. '
            f'<br><a href="{URL_APP_STREAMLIT}" target="_blank" '
            f'style="color:#2c5282; text-decoration:underline; font-weight:bold;">'
            f'\U0001f680 Abrir app desplegada \u2192</a>'
        )
        link7 = 'Abrir fase \u2192'
    else:
        descripcion_fase7 = 'Dashboard, vistas por perfil y documentación final.'
        link7 = 'Pendiente'
    H.append(f'                <a href="fase7/fase7_index.html" class="tarjeta" style="border-left-color:#2c5282{op7}">')
    H.append('                    <h3 style="color:#2c5282">\U0001f680 Fase 7: Aplicación</h3>')
    H.append(f'                    <p>{descripcion_fase7}</p>')
    H.append(f'                    <div class="link" style="color:#2c5282">{link7}</div>')
    H.append('                </a>')

    H.append('            </div>')
    H.append('        </section>')
    H.append('')

    # --- Sección AutoML ---
    H.append('        <section class="seccion">')
    H.append('            <h2>\u26a1 Pre-Modelado: AutoML</h2>')
    H.append('            <div style="background:#FFFBEB; padding:20px; border-radius:10px; border-left:4px solid #D69E2E; margin-bottom:15px;">')
    H.append('                <strong>\u26a0\ufe0f Requisito:</strong> Para usar AutoML, primero ejecuta las Fases 1, 2 y 3.')
    H.append('            </div>')
    H.append('            <a href="fase_automl/fase_automl_index.html"')
    H.append('               style="display:inline-block; padding:12px 24px; background:#5B21B6; color:white;')
    H.append('                      border-radius:8px; text-decoration:none; font-weight:bold;">')
    H.append('                \u26a1 Exploración de Frameworks AutoML \u2192')
    H.append('            </a>')
    H.append('        </section>')
    H.append('')

    # --- Sección Datos del Proyecto (con aviso instructivo si faltan Excel) ---
    H.append('        <section class="seccion">')
    H.append('            <h2>\U0001f4ca Datos del Proyecto</h2>')
    if excel_disponible:
        # Caso normal: Excel encontrados y leidos
        H.append('            <div style="background:#f8fafc; padding:20px; border-radius:10px; border-left:4px solid #3182ce; margin-bottom:20px;">')
        H.append(f'                <p><strong>\U0001f4c1 Origen:</strong> {n_archivos_excel} archivos Excel con {n_tablas} tablas</p>')
        H.append(f'                <p><strong>\U0001f4cb Tablas:</strong> {hojas_texto}</p>')
        H.append(f'                <p><strong>\U0001f4ca Registros totales (originales):</strong> {fmt_es_num(n_registros_raw)}</p>')
        H.append('            </div>')
    else:
        # Caso fallo: Excel no encontrados - mensaje rico instructivo
        H.append('            <div style="background:#fef3c7; padding:20px; border-radius:10px; border-left:4px solid #d97706; margin-bottom:20px;">')
        H.append('                <p><strong>\u26a0\ufe0f Excel originales no encontrados.</strong></p>')
        H.append('                <p>Para ejecutar este proyecto necesitas subir tus 2 archivos Excel a '
                 '<code>data/00_raw/</code> con esta estructura:</p>')
        H.append('                <ul style="margin-top:10px;">')
        H.append('                    <li><strong>datos_proyecto_sin_preinscrip.xlsx</strong> '
                 '— 8 hojas con nombres exactos: '
                 '<code>Titulaciones</code>, <code>Recibos</code>, <code>Domicilios</code>, '
                 '<code>Expedientes</code>, <code>Nac-Sexo_Nacionalidad</code>, '
                 '<code>Circunstancias</code>, <code>Trabajo</code>, <code>Notas</code></li>')
        H.append('                    <li><strong>preinscripcion_si.xlsx</strong> — 1 hoja con datos de preinscripción</li>')
        H.append('                </ul>')
        H.append('                <p style="margin-top:10px; font-size:0.9em; color:#92400e;">'
                 'Los nombres de los Excel se pueden cambiar en '
                 '<code>src/config_entorno.py</code> (constantes <code>EXCEL_PRINCIPAL</code> '
                 'y <code>EXCEL_PREINSCRIPCION</code>).</p>')
        H.append('            </div>')
    H.append('        </section>')
    H.append('')

    # --- Sección NUEVA: Métricas del Modelo Ganador (solo si JSON existe) ---
    if metricas_modelo:
        H.append('        <section class="seccion">')
        H.append('            <h2>\U0001f3c6 Métricas del Modelo Ganador</h2>')
        H.append('            <div style="background:#f0f9ff; padding:20px; border-radius:10px; '
                 'border-left:4px solid #1e4d8c; margin-bottom:20px;">')

        # --- Identidad del modelo ---
        # Patrón "if existe → mostrar / else → omitir" siguiendo la
        # convención del resto del fichero (no inventar valores).
        modelo_nombre     = metricas_modelo.get('modelo_nombre', 'No disponible')
        modelo_estrategia = metricas_modelo.get('modelo_estrategia', '')
        modelo_familia    = metricas_modelo.get('modelo_familia', '')
        modelo_desc       = metricas_modelo.get('modelo_descripcion', '')

        # Cuando estrategia != 'none', se incluye en el nombre entre paréntesis.
        # Si es 'none' (sin balanceo de clases), no añadir nada para evitar ruido.
        if modelo_estrategia and modelo_estrategia != 'none':
            nombre_a_mostrar = f'{modelo_nombre} ({modelo_estrategia})'
        else:
            nombre_a_mostrar = modelo_nombre

        H.append(f'                <p><strong>\U0001f916 Modelo:</strong> {nombre_a_mostrar}</p>')

        if modelo_familia:
            H.append(f'                <p style="margin-left:24px; color:#4a5568; font-size:0.95em;">'
                     f'<strong>Familia:</strong> {modelo_familia}</p>')

        if modelo_desc:
            H.append(f'                <p style="margin-left:24px; color:#718096; font-size:0.9em; '
                     f'font-style:italic;">{modelo_desc}</p>')

        # --- Métricas de rendimiento ---
        if metricas_modelo.get('f1') is not None:
            H.append(f'                <p><strong>\U0001f3af F1 (clase abandono):</strong> '
                     f'{fmt_es_dec(metricas_modelo["f1"], 4)}</p>')
        if metricas_modelo.get('auc') is not None:
            H.append(f'                <p><strong>\U0001f4c8 AUC-ROC:</strong> '
                     f'{fmt_es_dec(metricas_modelo["auc"], 4)}</p>')
        if metricas_modelo.get('precision') is not None:
            H.append(f'                <p><strong>\U0001f50d Precision:</strong> '
                     f'{fmt_es_dec(metricas_modelo["precision"], 4)}</p>')
        if metricas_modelo.get('recall') is not None:
            H.append(f'                <p><strong>\U0001f514 Recall:</strong> '
                     f'{fmt_es_dec(metricas_modelo["recall"], 4)}</p>')
        if metricas_modelo.get('accuracy') is not None:
            H.append(f'                <p><strong>\U00002705 Accuracy:</strong> '
                     f'{fmt_es_dec(metricas_modelo["accuracy"], 4)}</p>')

        # --- Línea final: contexto de evaluación ---
        n_test = metricas_modelo.get('n_test')
        n_feat = metricas_modelo.get('n_features')
        n_feat_t = metricas_modelo.get('n_features_tecnicas')
        if n_test is not None or n_feat is not None:
            partes_ctx = []
            if n_test is not None:
                partes_ctx.append(f'<strong>{fmt_es_num(n_test)}</strong> casos de test')
            if n_feat is not None and n_feat_t is not None and n_feat != n_feat_t:
                partes_ctx.append(
                    f'<strong>{n_feat}</strong> features ({n_feat_t} técnicas)'
                )
            elif n_feat is not None:
                partes_ctx.append(f'<strong>{n_feat}</strong> features')
            ctx_html = ' · '.join(partes_ctx)
            H.append(f'                <p style="margin-top:12px; padding-top:10px; '
                     f'border-top:1px solid #cbd5e0; color:#4a5568; font-size:0.9em;">'
                     f'Evaluado sobre {ctx_html}.</p>')

        H.append('            </div>')
        H.append('        </section>')
        H.append('')

    # --- Sección KPIs ---
    H.append('        <section class="seccion">')
    H.append('            <h2>\U0001f4c8 Estadísticas del Proyecto</h2>')
    H.append(f'            <div class="kpis">{kpis_gen}')
    H.append(f'            </div>{aviso}')
    H.append('        </section>')
    H.append('')
    H.append('    </div>')
    H.append('')

    # =================================================================
    # BLOQUE: Resumen del proyecto y orquestador (V2 — AU_UJI Dinámico)
    # =================================================================
    # Añade un acceso al HTML resumen del orquestador (orquestador_resumen.html)
    # generado por src/html/generar_resumen_proyecto.py.
    # El acceso aparece SIEMPRE, pero el HTML solo existe tras ejecutar:
    #   - notebooks/fase0_configuracion/orquestador_maestro.ipynb, o
    #   - notebooks/fase0_configuracion/f0_actualizar_resumen.ipynb
    # Si el HTML no existe aún, el enlace dará 404 — eso es señal clara
    # al usuario de que debe regenerarlo.
    # =================================================================
    resumen_existe = (RUTA_HTML / 'orquestador_resumen.html').exists()
    bg_resumen = '#ebf4ff' if resumen_existe else '#fffaf0'
    border_resumen = '#3182ce' if resumen_existe else '#ed8936'
    color_titulo = '#2c5282' if resumen_existe else '#7b341e'
    estado_resumen_txt = (
        'Resumen disponible — pulsa para ver estado actual del proyecto y última ejecución del orquestador.'
        if resumen_existe else
        'Aún no se ha generado el resumen. Ejecuta el orquestador maestro o el notebook <code>f0_actualizar_resumen.ipynb</code>.'
    )
    boton_html = (
        '<a href="orquestador_resumen.html" '
        'style="display:inline-block; padding:10px 18px; background:#3182ce; color:white; '
        'border-radius:6px; text-decoration:none; font-weight:600; margin-top:10px;">'
        '🎓 Ver resumen del proyecto</a>'
    ) if resumen_existe else (
        '<span style="display:inline-block; padding:10px 18px; background:#a0aec0; color:white; '
        'border-radius:6px; font-weight:600; margin-top:10px; opacity:0.7;">'
        '🎓 Resumen pendiente de generar</span>'
    )

    H.append('    <div class="container" style="margin-top:0;">')
    H.append('        <section class="seccion" '
             f'style="background:{bg_resumen}; border-left:4px solid {border_resumen}; '
             f'padding:18px 22px; border-radius:8px;">')
    H.append(f'            <h2 style="color:{color_titulo}; font-size:18px;">'
             '🎓 Resumen del proyecto y orquestador</h2>')
    H.append(f'            <p style="margin-top:6px; color:#4a5568;">{estado_resumen_txt}</p>')
    H.append(f'            {boton_html}')
    H.append('        </section>')
    H.append('    </div>')
    H.append('')

    H.append('    <footer>')
    H.append(f'        <strong>{AUTORA}</strong> |')
    H.append(f'        <a href="mailto:{EMAIL_UOC}">{EMAIL_UOC}</a> (UOC) |')
    H.append(f'        <a href="mailto:{EMAIL_UJI}">{EMAIL_UJI}</a> (UJI)')
    H.append('        <br>')
    H.append(f'        <a href="{GITHUB_REPO}" target="_blank">\U0001f4d3 GitHub</a>')
    H.append(f'        <br>Generado: {fecha}')
    H.append('    </footer>')
    H.append('</body>')
    H.append('</html>')

    html = '\n'.join(H)

    # --- Guardar ---
    ruta_index = RUTA_HTML / 'index.html'
    ruta_index.parent.mkdir(parents=True, exist_ok=True)
    ruta_index.write_text(html, encoding='utf-8')

    if verbose:
        estado_kpis = []
        if kpi_expedientes:
            estado_kpis.append(f'Expedientes: {fmt_es_num(kpi_expedientes)}')
        if kpi_periodo:
            estado_kpis.append(f'Período: {kpi_periodo}')
        if kpi_variables:
            estado_kpis.append(f'Variables: {kpi_variables}')
        if kpi_abandono is not None:
            estado_kpis.append(f'Abandono: {fmt_es_dec(kpi_abandono, 1)}%')
        if kpi_auc is not None:
            estado_kpis.append(f'AUC: {fmt_es_dec(kpi_auc, 4)}')

        if estado_kpis:
            print(f'  \u2705 KPIs: {" | ".join(estado_kpis)}')
        else:
            print('  \u26a0 KPIs pendientes (ejecuta las fases primero)')

        # Estado de fases dinámicas
        estado_fases = []
        emoji_ok    = '\u2705'  # ✅
        emoji_clock = '\u23f3'  # ⏳
        estado_fases.append(f'Fase 5: {emoji_ok if fase5_done else emoji_clock}')
        estado_fases.append(f'Fase 6: {emoji_ok if fase6_done else emoji_clock}')
        estado_fases.append(f'Fase 7: {emoji_ok if fase7_done else emoji_clock}')
        print(f'  \U0001f4ca Fases: {" | ".join(estado_fases)}')

        print(f'  \u2705 {ruta_index}')

    # --- Generar también fase7_index.html ---
    # (página de la app Streamlit, dinámica como esta)
    try:
        actualizar_fase7_index(verbose=verbose)
    except Exception as e:
        if verbose:
            print(f'  \u26a0 Error generando fase7_index.html: {e}')

    return ruta_index


# ============================================================================
# GENERACIÓN DEL HTML DE FASE 7 (página de la aplicación Streamlit)
# ============================================================================
# Genera dinámicamente docs/html/fase7/fase7_index.html con:
#   - Botón grande con la URL de la app desplegada (de URL_APP_STREAMLIT)
#   - 6 tarjetas describiendo las vistas de la app
#   - Sección de detalles técnicos
#
# Todo el contenido se construye leyendo constantes de src/config_*.py.
# Cero hardcodes: si cambia la URL, autora, emails, etc., basta con cambiar
# el config y re-ejecutar `actualizar_index()`.
#
# Se llama automáticamente desde `actualizar_index()` al final.
# ============================================================================

def actualizar_fase7_index(verbose=True):
    """
    Regenera docs/html/fase7/fase7_index.html dinámicamente.

    El HTML se construye usando las constantes de src.config_* (URL de la
    app, autora, emails, GitHub) y la navegación superior de las otras
    fases (vía src.html.navegacion). Cero datos hardcoded.

    Parameters
    ----------
    verbose : bool
        Si True, imprime progreso por consola.

    Returns
    -------
    Path
        Ruta del fase7_index.html generado.
    """
    from datetime import datetime as _dt

    # --- Imports de config (todo centralizado) ---
    from src.config import (
        BASE_PATH, RUTA_HTML,
        AUTORA, EMAIL_UOC, EMAIL_UJI, GITHUB_REPO,
        URL_APP_STREAMLIT, COLORES,
    )

    # --- Navegación superior (reusa la lógica de src.html.navegacion) ---
    try:
        from src.html.navegacion import generar_html_nav_fases
        nav_html = generar_html_nav_fases(fase_activa='fase7', ruta_base='..')
    except Exception as e:
        if verbose:
            print(f'  \u26a0 Error cargando navegacion.py: {e}')
        nav_html = ''  # fallback: sin nav superior

    fecha = _dt.now().strftime('%d/%m/%Y')
    assets_rel = '../../assets'

    # --- Vistas de la app (descripciones humanas) ---
    # Esto es información del proyecto (no resultados), por eso va en código.
    # Si añades una vista nueva, edita aquí.
    vistas_app = [
        {
            'icono':       '\U0001f3db\ufe0f',  # 🏛️
            'titulo':      'Vista Institucional',
            'descripcion': 'KPIs globales del proyecto y métricas del modelo ganador.',
            'color':       '#1e4d8c',
        },
        {
            'icono':       '\U0001f393',  # 🎓
            'titulo':      'Por Titulación',
            'descripcion': 'Análisis específico de cada titulación de la UJI.',
            'color':       '#3182ce',
        },
        {
            'icono':       '\U0001f331',  # 🌱
            'titulo':      'Futuro Estudiante',
            'descripcion': 'Pronóstico para alumnos que están considerando matricularse.',
            'color':       '#38a169',
        },
        {
            'icono':       '\U0001f4da',  # 📚
            'titulo':      'En Curso',
            'descripcion': 'Pronóstico para alumnos ya matriculados.',
            'color':       '#ed8936',
        },
        {
            'icono':       '\u2696\ufe0f',  # ⚖️
            'titulo':      'Equidad',
            'descripcion': 'Análisis de fairness por sexo, vía de acceso, edad, rama.',
            'color':       '#805ad5',
        },
        {
            'icono':       '\U0001f4d6',  # 📖
            'titulo':      'Leyenda',
            'descripcion': 'Glosario, paleta de colores y transparencia metodológica.',
            'color':       '#718096',
        },
    ]

    # --- Construcción del HTML ---
    H = []
    H.append('<!DOCTYPE html>')
    H.append('<html lang="es">')
    H.append('<head>')
    H.append('    <meta charset="UTF-8">')
    H.append('    <meta name="viewport" content="width=device-width, initial-scale=1.0">')
    H.append('    <title>Fase 7: Aplicación | TFM Abandono Universitario</title>')
    H.append('    <link rel="stylesheet" href="../style.css">')
    H.append('</head>')
    H.append('<body>')

    # --- HEADER ---
    H.append('    <header>')
    H.append('        <div class="header-content">')
    H.append(f'            <img src="{assets_rel}/logo_uoc.png" alt="UOC" class="header-logo" onerror="this.style.display=\'none\'">')
    H.append('            <div class="header-title">')
    H.append('                <h1>\U0001f680 Fase 7: Aplicación</h1>')
    H.append('                <p>Aplicación Streamlit del TFM</p>')
    H.append('            </div>')
    H.append(f'            <img src="{assets_rel}/logo_uji.jpg" alt="UJI" class="header-logo" onerror="this.style.display=\'none\'">')
    H.append('        </div>')
    H.append('    </header>')
    H.append('')

    # --- NAV SUPERIOR (de navegacion.py) ---
    H.append('    ' + nav_html.replace('\n', '\n    ') if nav_html else '')
    H.append('')

    H.append('    <div class="container">')
    H.append('')

    # --- SECCIÓN 1: BOTÓN GRANDE A LA APP ---
    H.append('        <section class="seccion">')
    H.append('            <h2>\U0001f680 Aplicación Streamlit</h2>')
    H.append('            <p>La aplicación de este TFM está desplegada y disponible online:</p>')
    H.append(f'            <a href="{URL_APP_STREAMLIT}" target="_blank"')
    H.append('               style="display:inline-block; padding:18px 36px; background:#2c5282;')
    H.append('                      color:white; border-radius:10px; text-decoration:none;')
    H.append('                      font-weight:bold; font-size:1.1rem; margin: 20px 0;">')
    H.append('                \U0001f680 Abrir app desplegada \u2192')
    H.append('            </a>')
    H.append('            <p style="color:#718096; font-size:0.9em; margin-top:10px;">')
    H.append('                Se abrirá en una nueva pestaña.')
    H.append('            </p>')
    H.append('        </section>')
    H.append('')

    # --- SECCIÓN 2: VISTAS DE LA APP ---
    H.append('        <section class="seccion">')
    H.append('            <h2>\U0001f4cb Vistas disponibles</h2>')
    H.append('            <p>La app está organizada en 6 vistas según el perfil de usuario:</p>')
    H.append('            <div class="grid-tarjetas">')
    for vista in vistas_app:
        H.append(f'                <div class="tarjeta" style="border-left-color:{vista["color"]}">')
        H.append(f'                    <h3 style="color:{vista["color"]}">{vista["icono"]} {vista["titulo"]}</h3>')
        H.append(f'                    <p>{vista["descripcion"]}</p>')
        H.append('                </div>')
    H.append('            </div>')
    H.append('        </section>')
    H.append('')

    # --- SECCIÓN 3: DETALLES TÉCNICOS ---
    url_visible = URL_APP_STREAMLIT.replace('https://', '').rstrip('/')
    H.append('        <section class="seccion">')
    H.append('            <h2>\U0001f6e0\ufe0f Detalles técnicos</h2>')
    H.append('            <div style="background:#f8fafc; padding:20px; border-radius:10px;')
    H.append('                        border-left:4px solid #3182ce; margin-bottom:20px;">')
    H.append('                <p><strong>\U0001f4e6 Framework:</strong> Streamlit</p>')
    H.append('                <p><strong>\u2601\ufe0f Despliegue:</strong> Streamlit Cloud</p>')
    H.append(f'                <p><strong>\U0001f517 URL:</strong> '
             f'<a href="{URL_APP_STREAMLIT}" target="_blank">{url_visible}</a></p>')
    H.append(f'                <p><strong>\U0001f4c2 Repositorio:</strong> '
             f'<a href="{GITHUB_REPO}" target="_blank">{GITHUB_REPO.replace("https://", "")}</a></p>')
    H.append('                <p><strong>\U0001f40d Entry point:</strong> <code>app/main.py</code></p>')
    H.append('            </div>')
    H.append('        </section>')
    H.append('')

    H.append('    </div>')
    H.append('')

    # --- FOOTER ---
    H.append('    <footer>')
    H.append(f'        <strong>{AUTORA}</strong> |')
    H.append(f'        <a href="mailto:{EMAIL_UOC}">{EMAIL_UOC}</a> (UOC) |')
    H.append(f'        <a href="mailto:{EMAIL_UJI}">{EMAIL_UJI}</a> (UJI)')
    H.append('        <br>')
    H.append(f'        <a href="{GITHUB_REPO}" target="_blank">\U0001f4d3 GitHub</a>')
    H.append(f'        <br>Generado: {fecha}')
    H.append('    </footer>')
    H.append('</body>')
    H.append('</html>')

    html = '\n'.join(H)

    # --- Guardar ---
    ruta_fase7 = RUTA_HTML / 'fase7' / 'fase7_index.html'
    ruta_fase7.parent.mkdir(parents=True, exist_ok=True)
    ruta_fase7.write_text(html, encoding='utf-8')

    if verbose:
        print(f'  \u2705 {ruta_fase7}')

    return ruta_fase7
