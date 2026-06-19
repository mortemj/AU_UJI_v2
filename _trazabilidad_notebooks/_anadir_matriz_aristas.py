# -*- coding: utf-8 -*-
"""Parsea los .mmd de grafos/ y anade una hoja 'matriz_aristas' al Excel
inventario_notebooks.xlsx para auditar el grafo. SOLO escribe en ese Excel."""
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
GRAFOS = BASE / "grafos"
XLSX = BASE / "inventario_notebooks.xlsx"

# operadores de enlace mermaid, de mas largo a mas corto
OP = re.compile(
    r'(<-\.\s*"[^"]*"\s*\.-'        # <-. "x" .-   (reverso, discontinuo)
    r'|-\.\s*"[^"]*"\s*\.->'        # -. "x" .->   (discontinuo etiquetado)
    r'|--\s*"[^"]*"\s*-->'          # -- "x" -->   (solido etiquetado)
    r'|-\.->'                       # -.->         (discontinuo)
    r'|-->'                         # -->          (solido)
    r'|<-\.'                        # <-.          (reverso)
    r'|\.-)'                        # .-           (cola discontinua)
)
NODE_DEF = re.compile(r'(\w+)\s*(\[\([^)]*?\)\]|\(\[[^\]]*?\]\)|\[[^\]]*?\]|\([^)]*?\)|\{[^}]*?\})')


def limpiar(txt):
    txt = txt.strip().strip('"').strip("()[]{}").strip('"')
    txt = txt.replace("<br/>", " ").replace("<br>", " ")
    return re.sub(r"\s+", " ", txt).strip()


def node_id(token):
    m = re.match(r'\s*(\w+)', token)
    return m.group(1) if m else token.strip()


def etiqueta_op(op):
    m = re.search(r'"([^"]*)"', op)
    return m.group(1) if m else ""


def tipo_op(op):
    return "solida" if ("-->" in op and "." not in op.split('"')[0] and "-. " not in op) and op.strip() in ('-->',) or op.strip().startswith("--") and "-->" in op else (
        "solida" if op.strip() == "-->" or op.strip().startswith('-- "') else "discontinua")


def es_solida(op):
    s = op.strip()
    return s == "-->" or s.startswith('-- "')


def reverso(op):
    return op.strip().startswith("<-.")


rows = []  # (grafo, origen_id, origen_label, sentido, tipo, etiqueta, destino_id, destino_label)
labels = {}

mmd_files = sorted(GRAFOS.glob("*.mmd"))
# primera pasada: etiquetas globales por fichero
for f in mmd_files:
    text = f.read_text(encoding="utf-8")
    file_labels = {}
    for m in NODE_DEF.finditer(text):
        file_labels[m.group(1)] = limpiar(m.group(2))
    labels[f.name] = file_labels

# segunda pasada: aristas
for f in mmd_files:
    grafo = f.stem
    file_labels = labels[f.name]
    for raw in f.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if line.startswith("%%") or not OP.search(line):
            continue
        # quitar comentario inline tras %% por si acaso
        line = line.split("%%")[0].strip()
        partes = OP.split(line)
        # partes = [nodo, op, nodo, op, nodo, ...]
        if len(partes) < 3:
            continue
        i = 0
        while i + 2 < len(partes):
            izq_tok, op, der_tok = partes[i], partes[i + 1], partes[i + 2]
            izq, der = node_id(izq_tok), node_id(der_tok)
            if not izq or not der:
                i += 2
                continue
            # etiqueta inline en token (define label) -> registrar
            for tok in (izq_tok, der_tok):
                m = NODE_DEF.match(tok.strip())
                if m and m.group(1) not in file_labels:
                    file_labels[m.group(1)] = limpiar(m.group(2))
            tipo = "solida" if es_solida(op) else "discontinua"
            if reverso(op):
                origen, destino = der, izq
            else:
                origen, destino = izq, der
            rows.append((
                grafo,
                origen, file_labels.get(origen, origen),
                "origen->destino",
                tipo,
                etiqueta_op(op),
                destino, file_labels.get(destino, destino),
            ))
            i += 2

# escribir hoja
wb = openpyxl.load_workbook(XLSX)
if "matriz_aristas" in wb.sheetnames:
    del wb["matriz_aristas"]
ws = wb.create_sheet("matriz_aristas")
cab = ["grafo", "origen_id", "origen_etiqueta", "sentido", "tipo_arista",
       "etiqueta_arista", "destino_id", "destino_etiqueta"]
ws.append(cab)
for r in rows:
    ws.append(list(r))

anchos = {"A": 14, "B": 14, "C": 42, "D": 16, "E": 13, "F": 40, "G": 14, "H": 42}
for col, w in anchos.items():
    ws.column_dimensions[col].width = w
head_fill = PatternFill("solid", fgColor="1F4E78")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = head_fill
    c.alignment = Alignment(vertical="center", horizontal="center")
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"

wb.save(XLSX)
print(f"OK -> hoja 'matriz_aristas' con {len(rows)} aristas")
sol = sum(1 for r in rows if r[4] == "solida")
print(f"  solidas={sol}  discontinuas={len(rows)-sol}")
from collections import Counter
for g, n in sorted(Counter(r[0] for r in rows).items()):
    s = sum(1 for r in rows if r[0] == g and r[4] == "solida")
    print(f"  {g}: {n} aristas ({s} solidas)")
