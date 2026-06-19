# -*- coding: utf-8 -*-
"""Auditoria a nivel de ARTEFACTO (fichero) a partir de la hoja 'inventario'.

Para cada fichero (parquet/pkl/json/csv/xlsx/html/png/pdf/...) detecta:
  - quien lo PRODUCE (aparece en su columna 'salidas')
  - quien lo CONSUME (aparece en su columna 'entradas')
y clasifica su estado, marcando posibles ENLACES ROTOS (entrada sin productor)
y SALIDAS NO CONSUMIDAS.

Heuristica: el emparejamiento se hace por NOMBRE DE FICHERO (basename), no por
ruta completa, porque el inventario mezcla rutas largas y cortas. Por eso un
mismo basename en dos carpetas (p.ej. dataset_final_tfm.parquet) se agrupa en una
sola fila y se listan todas las rutas vistas en la columna 'rutas'. Es una ayuda
de auditoria, no una verdad absoluta: revisar las filas marcadas.

SOLO escribe la hoja 'auditoria_artefactos' en inventario_notebooks.xlsx.
"""
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
XLSX = BASE / "inventario_notebooks.xlsx"

EXT = r"parquet|pkl|json|csv|xlsx|html|png|pdf|jpe?g|svg|zip"
RE_FICH = re.compile(r"([\w./\\{}<>*-]+\.(?:" + EXT + r"))", re.IGNORECASE)
# tipos que son entregables terminales esperados (no son 'enlace roto' si no se consumen)
TERMINAL = {"html", "png", "pdf", "jpg", "jpeg", "svg"}
# tipos de datos cuyo consumo sin productor SI es sospechoso
DATOS = {"parquet", "pkl", "json", "csv", "xlsx", "zip"}


def basename(tok):
    tok = tok.replace("\\", "/")
    return tok.rsplit("/", 1)[-1].lower()


def ext(name):
    return name.rsplit(".", 1)[-1].lower()


def expand_braces(s):
    """Expande llaves estilo shell: '{a,b}.x' -> ['a.x','b.x'] (soporta anidadas/secuenciales)."""
    m = re.search(r"\{([^{}]*)\}", s)
    if not m:
        return [s]
    pre, post = s[:m.start()], s[m.end():]
    out = []
    for o in m.group(1).split(","):
        out.extend(expand_braces(pre + o.strip() + post))
    return out


def es_patron(name):
    """Glob o placeholder: no es un fichero concreto auditable."""
    return ("*" in name) or ("<" in name) or (">" in name) or ("{" in name) or ("}" in name)


def ficheros(celda):
    if not celda:
        return []
    res = []
    for token in str(celda).split(";"):
        for expandido in expand_braces(token):
            res.extend(RE_FICH.findall(expandido))
    return res


wb = openpyxl.load_workbook(XLSX)
inv = wb["inventario"]
filas = list(inv.iter_rows(min_row=2, values_only=True))
# columnas: notebook, fase, que_hace, entradas, salidas, uso
NB, FASE, ENTR, SAL = 0, 1, 3, 4

productores = defaultdict(set)   # basename -> {notebooks}
consumidores = defaultdict(set)  # basename -> {notebooks}
rutas = defaultdict(set)         # basename -> {rutas completas vistas}

for row in filas:
    nb = row[NB]
    for f in ficheros(row[SAL]):
        b = basename(f)
        productores[b].add(nb)
        rutas[b].add(f.replace("\\", "/"))
    for f in ficheros(row[ENTR]):
        b = basename(f)
        consumidores[b].add(nb)
        rutas[b].add(f.replace("\\", "/"))

todos = sorted(set(productores) | set(consumidores))
rows_out = []
for b in todos:
    prod = sorted(productores.get(b, []))
    cons = sorted(consumidores.get(b, []))
    e = ext(b)
    rset = sorted(rutas[b])
    multi_ruta = "SI" if len({r.rsplit("/", 1)[0] for r in rset if "/" in r}) > 1 else ""
    if es_patron(b):
        estado = "patron glob/placeholder (no auditable)"
    elif e in TERMINAL:
        estado = "terminal (web/figura)"
    elif prod and cons:
        estado = "ok (producido y consumido)"
    elif prod and not cons:
        estado = "salida no consumida"
    elif cons and not prod:
        estado = "ENLACE ROTO: entrada sin productor" if e in DATOS else "entrada sin productor"
    else:
        estado = "?"
    rows_out.append((
        b, e, estado,
        "; ".join(prod) if prod else "(nadie)",
        "; ".join(cons) if cons else "(nadie)",
        "; ".join(rset),
        multi_ruta,
    ))

# orden: primero enlaces rotos, luego salidas no consumidas de datos, resto
def clave(r):
    pr = {"ENLACE ROTO: entrada sin productor": 0, "salida no consumida": 1,
          "entrada sin productor": 2, "ok (producido y consumido)": 3,
          "terminal (web/figura)": 4, "patron glob/placeholder (no auditable)": 6}.get(r[2], 5)
    # salidas no consumidas de datos antes que las terminales ya estan separadas
    return (pr, r[0])

rows_out.sort(key=clave)

if "auditoria_artefactos" in wb.sheetnames:
    del wb["auditoria_artefactos"]
ws = wb.create_sheet("auditoria_artefactos")
cab = ["artefacto", "tipo", "estado", "producido_por", "consumido_por", "rutas_vistas", "multi_ruta"]
ws.append(cab)
for r in rows_out:
    ws.append(list(r))

anchos = {"A": 36, "B": 9, "C": 34, "D": 40, "E": 40, "F": 50, "G": 10}
for col, w in anchos.items():
    ws.column_dimensions[col].width = w
head_fill = PatternFill("solid", fgColor="1F4E78")
roto_fill = PatternFill("solid", fgColor="F8CBAD")
nocons_fill = PatternFill("solid", fgColor="FFF2CC")
for c in ws[1]:
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = head_fill
    c.alignment = Alignment(vertical="center", horizontal="center")
for i, r in enumerate(rows_out, start=2):
    for c in ws[i]:
        c.alignment = Alignment(vertical="top", wrap_text=True)
    if r[2].startswith("ENLACE ROTO"):
        for c in ws[i]:
            c.fill = roto_fill
    elif r[2] == "salida no consumida" and r[1] in DATOS:
        for c in ws[i]:
            c.fill = nocons_fill
ws.freeze_panes = "A2"
wb.save(XLSX)

# resumen consola
from collections import Counter
cnt = Counter(r[2] for r in rows_out)
print(f"OK -> hoja 'auditoria_artefactos' con {len(rows_out)} artefactos")
for k, v in sorted(cnt.items()):
    print(f"  {v:3d}  {k}")
print("\nENLACES ROTOS (datos consumidos sin productor en el inventario):")
rotos = [r for r in rows_out if r[2].startswith("ENLACE ROTO")]
for r in rotos:
    print(f"  - {r[0]}  <- consumido por: {r[4]}")
print(f"\nSALIDAS DE DATOS NO CONSUMIDAS (posibles huerfanos de datos):")
for r in rows_out:
    if r[2] == "salida no consumida" and r[1] in DATOS:
        print(f"  - {r[0]}  (produce: {r[3]})")
print(f"\nArtefactos con MULTIPLES rutas (posible duplicacion de carpeta):")
for r in rows_out:
    if r[6] == "SI":
        print(f"  - {r[0]}  rutas: {r[5]}")
